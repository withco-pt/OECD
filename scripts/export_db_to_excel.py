#!/usr/bin/env python3
"""
Exportação da base de dados (Supabase) para um Excel com uma folha por entidade.

Lê as views/tabelas públicas via PostgREST (chave anon, só leitura — a mesma
usada pelo frontend) e gera um .xlsx com:
  - Resumo: nº de serviços e medições por entidade
  - Uma folha por entidade, com todas as medições (measurements_catalog)
  - Serviços: catálogo completo de serviços
  - Indicadores: catálogo completo de indicadores

Uso:
    python3 scripts/export_db_to_excel.py [caminho-saida.xlsx]

Sem argumento, grava em "Export-BD-por-entidade-<data-de-hoje>.xlsx" na raiz do
projeto (mesmo padrão dos exports anteriores).
"""
import json
import re
import sys
import urllib.request
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
ENV_FILE = ROOT / ".env.local"

TITLE_COLOR = "002B82"
SUBTITLE_COLOR = "64718B"
HEADER_FILL = "002B82"
HEADER_FONT_COLOR = "FFFFFF"

# Ordem e nomes de folha das entidades — nomes completos truncados a 31
# caracteres (limite do Excel), preservando o final quando possível.
ENTITY_ORDER = ["ec", "at", "iss", "cml", "adc"]
SHEET_NAME_OVERRIDES = {
    "at": "Autoridade Tributária e Aduanei",
    "adc": "Agência para o Desenv. e Coesão",
}

MONTHS_PT = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

# Ordem de tipo dentro de cada dimensão, para a folha "Indicadores"
# (aproxima a ordem observada nos exports anteriores).
TYPE_ORDER = {"user_experience": 0, "operational": 1, "compliance": 2}


def read_env():
    env = {}
    for line in ENV_FILE.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip()
    return env


def fetch_all(base_url, api_key, table, select, order=None, filter_qs=None, page_size=1000):
    """GET paginado a uma tabela/view PostgREST, devolve lista de dicts."""
    rows = []
    offset = 0
    while True:
        qs = f"select={select}"
        if order:
            qs += f"&order={order}"
        if filter_qs:
            qs += f"&{filter_qs}"
        url = f"{base_url}/rest/v1/{table}?{qs}"
        req = urllib.request.Request(url)
        req.add_header("apikey", api_key)
        req.add_header("Authorization", f"Bearer {api_key}")
        req.add_header("Range-Unit", "items")
        req.add_header("Range", f"{offset}-{offset + page_size - 1}")
        with urllib.request.urlopen(req) as resp:
            batch = json.loads(resp.read())
        rows.extend(batch)
        if len(batch) < page_size:
            break
        offset += page_size
    return rows


def sheet_name_for(short_name, full_name):
    if short_name in SHEET_NAME_OVERRIDES:
        return SHEET_NAME_OVERRIDES[short_name]
    return full_name[:31]


def style_title(cell, text):
    cell.value = text
    cell.font = Font(bold=True, size=14, color=TITLE_COLOR)


def style_subtitle(cell, text):
    cell.value = text
    cell.font = Font(size=9, color=SUBTITLE_COLOR)


def style_header_row(ws, row_idx, headers):
    fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    font = Font(bold=True, size=10, color=HEADER_FONT_COLOR)
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=text)
        cell.font = font
        cell.fill = fill


def sim_nao(v):
    return "Sim" if v else "Não"


def main():
    out_path = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / f"Export-BD-por-entidade-{date.today().isoformat()}.xlsx"

    env = read_env()
    base_url = env["NEXT_PUBLIC_SUPABASE_URL"]
    api_key = env["NEXT_PUBLIC_SUPABASE_ANON_KEY"]

    print("A carregar organizações…")
    orgs = fetch_all(base_url, api_key, "organizations", "short_name,name,area_governamental", order="short_name")
    org_by_short = {o["short_name"]: o for o in orgs}

    print("A carregar dimensões (thematic_priorities)…")
    priorities = fetch_all(base_url, api_key, "thematic_priorities", "id,name_pt,display_order", order="display_order")
    priority_order = {p["id"]: p["display_order"] for p in priorities}
    priority_name = {p["id"]: p["name_pt"] for p in priorities}

    print("A carregar indicadores…")
    indicators = fetch_all(
        base_url,
        api_key,
        "indicators",
        "id,description,thematic_priority_id,type_of_indicator,value_type,value_scale_min,value_scale_max,channel_scope,is_mandatory",
    )
    ind_by_id = {i["id"]: i for i in indicators}

    print("A carregar catálogo de serviços…")
    services = fetch_all(
        base_url,
        api_key,
        "services_catalog",
        "id,name,entity,entity_short,matriz_adotada,active,has_measurements",
        order="entity,name",
    )
    service_name_by_id = {s["id"]: s["name"] for s in services}

    print("A carregar medições (measurements_catalog) — pode demorar…")
    measurements = fetch_all(
        base_url,
        api_key,
        "measurements_catalog",
        "entity_short,service_id,indicator_id,year,month,channel,geo_level,geo_name,value,value_text,category_counts,total_respondentes,total_inquiridos",
    )
    print(f"  {len(measurements)} medições carregadas.")

    measurements_by_entity = {}
    for m in measurements:
        measurements_by_entity.setdefault(m["entity_short"], []).append(m)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Resumo ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("Resumo")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    style_title(ws["A1"], "Exportação da Base de Dados — Matriz para a Inovação nos Serviços Públicos")
    style_subtitle(ws["A2"], f"OCDE/OPSI · dados extraídos em {date.today().isoformat()} (Supabase)")
    style_header_row(ws, 4, ["Entidade", "Área Governamental", "Nº Serviços", "Nº Medições"])

    row_idx = 5
    total_services = 0
    total_measurements = 0
    for short in ENTITY_ORDER:
        org = org_by_short.get(short)
        if not org:
            continue
        n_services = sum(1 for s in services if s["entity_short"] == short)
        n_measurements = len(measurements_by_entity.get(short, []))
        total_services += n_services
        total_measurements += n_measurements
        ws.cell(row=row_idx, column=1, value=org["name"])
        ws.cell(row=row_idx, column=2, value=org.get("area_governamental"))
        ws.cell(row=row_idx, column=3, value=n_services)
        ws.cell(row=row_idx, column=4, value=n_measurements)
        row_idx += 1

    bold = Font(bold=True)
    ws.cell(row=row_idx, column=1, value="TOTAL").font = bold
    ws.cell(row=row_idx, column=3, value=total_services).font = bold
    ws.cell(row=row_idx, column=4, value=total_measurements).font = bold

    # ── Uma folha por entidade ───────────────────────────────────────────
    MEAS_HEADERS = [
        "Serviço", "Dimensão", "Indicador", "Tipo", "Ano", "Mês", "Canal",
        "Nível Geo", "Geografia", "Valor", "Valor (texto)", "Sim", "Não",
        "Nº Respondentes", "Nº Inquiridos",
    ]
    MEAS_WIDTHS = [30, 24, 50, 14, 6, 7, 16, 10, 16, 10, 28, 7, 7, 14, 13]

    for short in ENTITY_ORDER:
        org = org_by_short.get(short)
        if not org:
            continue
        rows = measurements_by_entity.get(short, [])

        def sort_key(m):
            ind = ind_by_id.get(m["indicator_id"], {})
            return (
                service_name_by_id.get(m["service_id"], ""),
                priority_order.get(ind.get("thematic_priority_id"), 99),
                TYPE_ORDER.get(ind.get("type_of_indicator"), 9),
                ind.get("description", ""),
                m["year"],
                m["month"] or 0,
                m["channel"] or "",
                m["geo_name"] or "",
            )

        rows_sorted = sorted(rows, key=sort_key)

        ws = wb.create_sheet(sheet_name_for(short, org["name"]))
        for col_letter, width in zip("ABCDEFGHIJKLMNO", MEAS_WIDTHS):
            ws.column_dimensions[col_letter].width = width
        style_title(ws["A1"], org["name"])
        style_subtitle(ws["A2"], f"Medições — {org['name']}")
        style_header_row(ws, 4, MEAS_HEADERS)
        ws.freeze_panes = "A5"

        r = 5
        for m in rows_sorted:
            ind = ind_by_id.get(m["indicator_id"], {})
            cats = m.get("category_counts") or {}
            month_label = MONTHS_PT[m["month"] - 1] if m["month"] else None
            values = [
                service_name_by_id.get(m["service_id"], "(desconhecido)"),
                priority_name.get(ind.get("thematic_priority_id"), ""),
                ind.get("description", ""),
                ind.get("type_of_indicator", ""),
                m["year"],
                month_label,
                m["channel"],
                m["geo_level"],
                m["geo_name"],
                float(m["value"]) if m["value"] is not None else None,
                m["value_text"],
                cats.get("Sim"),
                cats.get("Não"),
                m["total_respondentes"],
                m["total_inquiridos"],
            ]
            for c, v in enumerate(values, start=1):
                ws.cell(row=r, column=c, value=v)
            r += 1

    # ── Serviços ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Serviços")
    for col_letter, width in zip("ABCDE", [34, 40, 15, 8, 14]):
        ws.column_dimensions[col_letter].width = width
    style_title(ws["A1"], "Catálogo de Serviços")
    style_header_row(ws, 3, ["Entidade", "Serviço", "Matriz Adotada", "Ativo", "Tem Medições"])
    ws.freeze_panes = "A4"

    services_sorted = sorted(services, key=lambda s: (s["entity"], s["name"]))
    r = 4
    for s in services_sorted:
        ws.cell(row=r, column=1, value=s["entity"])
        ws.cell(row=r, column=2, value=s["name"])
        ws.cell(row=r, column=3, value=sim_nao(s["matriz_adotada"]))
        ws.cell(row=r, column=4, value=sim_nao(s["active"]))
        ws.cell(row=r, column=5, value=sim_nao(s["has_measurements"]))
        r += 1

    # ── Indicadores ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Indicadores")
    for col_letter, width in zip("ABCDEFGH", [24, 52, 14, 20, 10, 10, 26, 12]):
        ws.column_dimensions[col_letter].width = width
    style_title(ws["A1"], "Catálogo de Indicadores")
    style_header_row(
        ws, 3,
        ["Dimensão", "Indicador", "Tipo", "Tipo de Valor", "Escala Min", "Escala Max", "Canal", "Obrigatório"],
    )
    ws.freeze_panes = "A4"

    def ind_sort_key(i):
        return (
            priority_order.get(i.get("thematic_priority_id"), 99),
            TYPE_ORDER.get(i.get("type_of_indicator"), 9),
            i.get("description", ""),
        )

    indicators_sorted = sorted(indicators, key=ind_sort_key)
    r = 4
    for i in indicators_sorted:
        ws.cell(row=r, column=1, value=priority_name.get(i.get("thematic_priority_id"), ""))
        ws.cell(row=r, column=2, value=i["description"])
        ws.cell(row=r, column=3, value=i["type_of_indicator"])
        ws.cell(row=r, column=4, value=i["value_type"])
        ws.cell(row=r, column=5, value=i.get("value_scale_min"))
        ws.cell(row=r, column=6, value=i.get("value_scale_max"))
        ws.cell(row=r, column=7, value=i.get("channel_scope"))
        ws.cell(row=r, column=8, value=sim_nao(i.get("is_mandatory")))
        r += 1

    wb.save(out_path)
    print(f"\nExportação concluída: {out_path}")
    print(f"  {total_services} serviços, {total_measurements} medições, {len(indicators)} indicadores.")


if __name__ == "__main__":
    main()
