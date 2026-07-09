#!/usr/bin/env python3
"""
Ingestão do questionário de experiência do utilizador (Matriz + LC) → measurements.

Objetivo: converter as respostas individuais dos ficheiros Excel do inquérito em
medições agregadas por (entidade, serviço, período, canal, indicador), preservando
fielmente a distribuição Sim/Não (que a amostra anterior tinha achatado para 1.0).

Uso:
    python3 scripts/ingest_survey.py --report        # dry-run: só valida e imprime
    python3 scripts/ingest_survey.py --sql out.sql   # gera SQL de ingestão

NADA é escrito na base de dados por este script — ou imprime relatório, ou gera SQL
para ser aplicado como migration (auditável e repetível para dados futuros).

O mapeamento coluna→indicador e texto→código está TODO explícito abaixo, para ser
revisto. Qualquer rótulo não previsto é reportado como UNMAPPED — nunca adivinhado.
"""
import argparse, glob, json, sys, unicodedata
from collections import defaultdict, Counter
from datetime import datetime

import openpyxl

BASE = "/Users/davidcruz/Library/CloudStorage/GoogleDrive-davidbernardocruz@gmail.com/Shared drives/WithOCDE_Prototype/02. Data/Data 01/"

# ── Normalização de texto ────────────────────────────────────────────────────
def norm(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return " ".join(s.split())

# ── Escalas texto → código 1-5 (por família de rótulos) ──────────────────────
FACILIDADE = {"muito dificil": 1, "dificil": 2, "nem facil nem dificil": 3, "neutro": 3, "facil": 4, "muito facil": 5}
RAPIDEZ = {"muito lento": 1, "lento": 2, "nem lento nem rapido": 3, "rapido": 4, "muito rapido": 5}
CONCORDANCIA = {"discordo totalmente": 1, "discordo parcialmente": 2, "nao concordo nem discordo": 3, "concordo parcialmente": 4, "concordo totalmente": 5}
CONCORDANCIA_DUR = {"discordo totalmente": 1, "discordo": 2, "nao concordo nem discordo": 3, "concordo": 4, "concordo totalmente": 5}
QUALIDADE = {"muito mau": 1, "mau": 2, "nem mau nem bom": 3, "bom": 4, "muito bom": 5}
SATISFACAO = {"muito insatisfatorio": 1, "insatisfatorio": 2, "nem insatisfatorio nem satisfatorio": 3, "satisfatorio": 4, "muito satisfatorio": 5}
ESPERA = {"muito demorado/inaceitavel": 1, "demorado": 2, "tempo aceitavel": 3, "rapido": 4, "muito rapido": 5}
EXPECTATIVAS = {
    "nao correspondeu nada as expectativas": 1,
    "nao correspondeu em parte nem totalmente as expectativas": 2,
    "correspondeu em parte as expectativas": 3,
    "correspondeu totalmente as expectativas": 4,
    "superou as expectativas": 5,
}
# Rótulos que NÃO contam (não respondeu / não aplicável ao indicador)
EXCLUDE = {"nao utilizei este canal", "nao se aplica", "nao se aplica.", "prefiro nao responder", "ainda aguardo conclusao/resposta", "neutro"}
# nota: "neutro" está em FACILIDADE como 3; removê-lo do exclude
EXCLUDE.discard("neutro")

# ── Mapa coluna Excel (1-indexed) → indicador ────────────────────────────────
# kind: likert | scale10 | nps | sim_nao | agendamento
# channel: fixo para as colunas do matriz de canais (17-23)
COLMAP = [
    (14, "ux_resolved",        "sim_nao",     None, None),
    (16, "ux_speed",           "likert",      RAPIDEZ, None),
    (17, "ux_channel_ease",    "likert",      FACILIDADE, "Presencial"),
    (18, "ux_channel_ease",    "likert",      FACILIDADE, "Telefone"),
    (19, "ux_channel_ease",    "likert",      FACILIDADE, "Digital/Online"),
    (20, "ux_channel_ease",    "likert",      FACILIDADE, "App"),
    (21, "ux_channel_ease",    "likert",      FACILIDADE, "Videochamada"),
    (22, "ux_channel_ease",    "likert",      FACILIDADE, "Chatbox"),
    (23, "ux_channel_ease",    "likert",      FACILIDADE, "Outro"),
    (24, "ux_findability",     "likert",      FACILIDADE, None),
    (25, "ux_language_simple", "likert",      CONCORDANCIA, None),
    (26, "ux_language_clear",  "likert",      CONCORDANCIA, None),
    (27, "ux_response_clarity","sim_nao",     None, None),
    (31, "ux_ease",            "likert",      FACILIDADE, None),
    (32, "ux_scheduling",      "agendamento", None, None),
    (33, "ux_scheduling_ease", "likert",      FACILIDADE, None),
    (34, "ux_scheduling_date", "likert",      FACILIDADE, None),
    (36, "ux_wait_time",       "likert",      ESPERA, None),
    (37, "ux_duration",        "likert",      CONCORDANCIA_DUR, None),
    (38, "ux_clarity_info",    "likert",      QUALIDADE, None),
    (39, "ux_knowledge",       "likert",      QUALIDADE, None),
    (40, "ux_routing",         "likert",      QUALIDADE, None),
    (41, "ux_fairness",        "likert",      CONCORDANCIA, None),
    (42, "ux_courtesy",        "likert",      SATISFACAO, None),
    (43, "ux_csat",            "scale10",     None, None),
    (44, "ux_expectations",    "likert",      EXPECTATIVAS, None),
    (45, "ux_nps",             "nps",         None, None),
]
# Colunas Sim/Não/N.A. ignoradas de propósito (sub-dimensões de transparência sem
# indicador dedicado no catálogo): 28 (Compreender a decisão), 29 (Saber o estado),
# 30 (Compreender os próximos passos). Reportadas mas não ingeridas.

# Coluna do Distrito/Região de Residência — confirmada na col 52 em ambos os
# ficheiros (Matriz: "Distrito de Residência"; LC: "Distrito ou Região Autónoma
# de Residência"). Gera uma segunda agregação (por distrito, sem canal) em
# paralelo à agregação por canal já existente — não substitui nem altera.
DISTRITO_COL = 52

SIM = "Sim"; NAO = "Não"; NA = "Não aplicável"; NMT = "Não, mas tentei"

# ── Identidade das organizações (public.organizations.id) ────────────────────
ORG_ID = {
    "at":  "08d2e2a3-65ca-4e35-8453-493f933bfe92",
    "ec":  "b8543b22-3400-4d20-9541-b756782fa31f",
    "iss": "dca9b165-c7d9-44b7-afd8-38ebc13e604c",
}
SCHEMA = {"at": "org_at", "ec": "org_ec", "iss": "org_iss"}

def entity_of(col8):
    n = norm(col8)
    if "tribut" in n or n == "at":
        return "at"
    if "seguranca" in n or "iss" in n:
        return "iss"
    if "espaco" in n or "cidadao" in n:
        return "ec"
    return None

def service_of(row, ent):
    # col 9 AT, col 10 ISS, col 11 EC, col 12 genérico
    idx = {"at": 9, "iss": 10, "ec": 11}.get(ent)
    v = row[idx - 1] if idx else None
    if not v:
        v = row[11]  # col 12 fallback
    return str(v).strip() if v else None

def period_of(row):
    v = row[2]  # Completion time (col 3)
    if isinstance(v, datetime):
        return v.year, v.month
    try:
        dt = datetime.fromisoformat(str(v))
        return dt.year, dt.month
    except Exception:
        for fmt in ("%m/%d/%y %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%m/%d/%Y %H:%M"):
            try:
                dt = datetime.strptime(str(v), fmt); return dt.year, dt.month
            except Exception:
                pass
    return None, None

def load():
    """Lê os ficheiros e devolve estruturas agregadas. Chave por NOME NORMALIZADO
    (não pelo texto cru) para não duplicar serviços por espaços/maiúsculas."""
    files = sorted(glob.glob(BASE + "Matriz*Question*.xlsx")) + sorted(glob.glob(BASE + "LC*Question*.xlsx"))
    assert files, "ficheiros não encontrados"

    agg = defaultdict(lambda: {"codes": [], "cats": Counter()})  # key=(ent,sn,y,m,chan,etl,kind)
    agg_geo = defaultdict(lambda: {"codes": [], "cats": Counter()})  # key=(ent,sn,y,m,geo_name,etl,kind)
    pop = Counter()                      # (ent, sn, y, m) -> nº respostas
    geo_pop = Counter()                  # distrito -> nº respostas (para relatório)
    display = defaultdict(Counter)       # (ent, sn) -> Counter(nome cru)
    unmapped = defaultdict(Counter)
    label_map = defaultdict(dict)
    skipped = Counter()

    for path in files:
        ws = openpyxl.load_workbook(path, data_only=True).active
        for r in range(2, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if all(v in (None, "") for v in row):
                continue
            ent = entity_of(row[7])
            if not ent:
                skipped["entidade_desconhecida"] += 1; continue
            serv = service_of(row, ent)
            if not serv:
                skipped["servico_em_branco"] += 1; continue
            y, m = period_of(row)
            if not y:
                skipped["periodo_invalido"] += 1; continue
            sn = norm(serv)
            pop[(ent, sn, y, m)] += 1
            display[(ent, sn)][serv.strip()] += 1

            distrito_raw = row[DISTRITO_COL - 1]
            geo_name = str(distrito_raw).strip() if distrito_raw not in (None, "") else None
            if geo_name:
                geo_pop[geo_name] += 1
            else:
                skipped["distrito_em_branco"] += 1

            for col, etl, kind, scale, chan in COLMAP:
                raw = row[col - 1]
                if raw in (None, ""):
                    continue
                key = (ent, sn, y, m, chan, etl, kind)
                # Segunda agregação, em paralelo: por distrito, sem canal (respondente
                # não é dividido por canal E distrito ao mesmo tempo — amostras já pequenas).
                geo_key = (ent, sn, y, m, geo_name, etl, kind) if geo_name else None
                if kind in ("sim_nao", "agendamento"):
                    nv = norm(raw)
                    cat = None
                    if nv == "sim":
                        cat = SIM
                    elif nv == "nao":
                        cat = NAO
                    elif nv == "nao aplicavel":
                        cat = NA
                    elif nv in ("nao, mas tentei", "nao mas tentei"):
                        cat = NMT
                    if cat is None:
                        unmapped[etl][str(raw).strip()] += 1
                    else:
                        agg[key]["cats"][cat] += 1
                        if geo_key:
                            agg_geo[geo_key]["cats"][cat] += 1
                elif kind == "likert":
                    nv = norm(raw)
                    if nv in EXCLUDE:
                        continue
                    code = scale.get(nv)
                    if code is None:
                        unmapped[etl][str(raw).strip()] += 1
                    else:
                        agg[key]["codes"].append(code)
                        label_map[etl][str(raw).strip()] = code
                        if geo_key:
                            agg_geo[geo_key]["codes"].append(code)
                elif kind in ("scale10", "nps"):
                    try:
                        fv = float(raw)
                    except Exception:
                        unmapped[etl][str(raw).strip()] += 1
                    else:
                        agg[key]["codes"].append(fv)
                        if geo_key:
                            agg_geo[geo_key]["codes"].append(fv)
    return dict(agg), dict(agg_geo), pop, geo_pop, display, unmapped, label_map, skipped, files


def compute(kind, d):
    """Devolve (value, category_counts|None, respondentes) para um grupo."""
    codes, cats = d["codes"], d["cats"]
    if kind == "likert":
        n = len(codes)
        return (round(sum(codes) / n, 2) if n else None), None, n
    if kind == "scale10":
        n = len(codes)
        return (round(sum(codes) / n, 2) if n else None), None, n
    if kind == "nps":
        n = len(codes)
        if not n:
            return None, None, 0
        promo = sum(1 for v in codes if v >= 9)
        detr = sum(1 for v in codes if v <= 6)
        pas = n - promo - detr
        nps = round(100 * (promo - detr) / n, 1)
        return nps, {"Promotores": promo, "Passivos": pas, "Detratores": detr}, n
    if kind == "sim_nao":
        s, nn, na = cats[SIM], cats[NAO], cats[NA]
        den = s + nn
        val = round(100 * s / den, 1) if den else None      # taxa de Sim (%)
        cc = {SIM: s, NAO: nn}                                # chaves base sempre presentes
        if na:
            cc[NA] = na
        return val, cc, s + nn + na
    if kind == "agendamento":
        s, nn, nmt = cats[SIM], cats[NAO], cats[NMT]
        return None, {SIM: s, NAO: nn, NMT: nmt}, s + nn + nmt  # sem ordem → value null
    return None, None, 0


def q(s):
    return "'" + str(s).replace("'", "''") + "'"


def emit_sql(out):
    agg, agg_geo, pop, geo_pop, display, unmapped, label_map, skipped, files = load()
    assert not unmapped, f"Há rótulos não mapeados, abortar: {dict(unmapped)}"

    # serviços distintos por (ent, sn) com nome de apresentação canónico
    svc_display = {k: c.most_common(1)[0][0] for k, c in display.items()}

    L = []
    L.append("-- Gerado por scripts/ingest_survey.py — NÃO editar à mão.")
    L.append(f"-- Fontes: {', '.join(p.split('/')[-1] for p in files)}")
    L.append("-- Substitui as medições de amostra pelas agregações reais do questionário.\n")

    # 1) Criar serviços em falta (por entidade), idempotente
    L.append("-- 1) Serviços em falta")
    for (ent, sn), disp in sorted(svc_display.items()):
        sch, org = SCHEMA[ent], ORG_ID[ent]
        L.append(
            f"INSERT INTO {sch}.services (name, name_normalized, organization_id, matriz_adotada, active) "
            f"SELECT {q(disp)}, {q(sn)}, {q(org)}::uuid, false, true "
            f"WHERE NOT EXISTS (SELECT 1 FROM {sch}.services WHERE name_normalized = {q(sn)});"
        )

    # 2) Limpar medições existentes (todas de questionário/amostra)
    L.append("\n-- 2) Limpar medições de amostra")
    for ent in ("at", "ec", "iss"):
        L.append(f"DELETE FROM {SCHEMA[ent]}.measurements;")

    # 3) Inserir agregações — 1 statement por entidade (VALUES + JOIN aos ids)
    # Duas famílias de linhas, em paralelo (não se substituem):
    #   - por canal (chan preenchido, geo_level/geo_name NULL) — comportamento existente
    #   - por distrito (geo_level='distrito', geo_name preenchido, channel NULL) — novo
    L.append("\n-- 3) Inserir medições reais")
    rows_by_ent = defaultdict(list)
    for (ent, sn, y, m, chan, etl, kind), d in sorted(agg.items(), key=lambda kv: tuple(str(x) for x in kv[0])):
        value, cc, resp = compute(kind, d)
        inq = pop[(ent, sn, y, m)]
        vstr = "NULL" if value is None else repr(float(value))
        cstr = "NULL" if cc is None else q(json.dumps(cc, ensure_ascii=False))
        chstr = "NULL" if chan is None else q(chan)
        rows_by_ent[ent].append(f"({q(sn)},{q(etl)},{y},{m},{chstr},NULL,NULL,{vstr},{cstr},{resp},{inq})")
    for (ent, sn, y, m, geo_name, etl, kind), d in sorted(agg_geo.items(), key=lambda kv: tuple(str(x) for x in kv[0])):
        value, cc, resp = compute(kind, d)
        inq = pop[(ent, sn, y, m)]
        vstr = "NULL" if value is None else repr(float(value))
        cstr = "NULL" if cc is None else q(json.dumps(cc, ensure_ascii=False))
        rows_by_ent[ent].append(f"({q(sn)},{q(etl)},{y},{m},NULL,{q('distrito')},{q(geo_name)},{vstr},{cstr},{resp},{inq})")

    nrows = 0
    for ent in ("at", "ec", "iss"):
        rows = rows_by_ent[ent]
        if not rows:
            continue
        nrows += len(rows)
        sch = SCHEMA[ent]
        L.append(
            f"INSERT INTO {sch}.measurements "
            f"(service_id, indicator_id, year, month, channel, geo_level, geo_name, value, category_counts, "
            f"total_respondentes, total_inquiridos, source_file)\nSELECT s.id, i.id, v.year, v.month, "
            f"v.channel, v.geo_level, v.geo_name, v.value::numeric, v.cc::jsonb, v.resp, v.inq, 'questionario_2026'\n"
            f"FROM (VALUES\n  " + ",\n  ".join(rows) +
            f"\n) AS v(sn, etl, year, month, channel, geo_level, geo_name, value, cc, resp, inq)\n"
            f"JOIN {sch}.services s ON s.name_normalized = v.sn\n"
            f"JOIN public.indicators i ON i.etl_column_key = v.etl;"
        )

    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(L) + "\n")
    print(f"SQL escrito em {out}: {len(svc_display)} serviços, {nrows} medições.")


def report():
    agg, agg_geo, pop, geo_pop, display, unmapped, label_map, skipped, files = load()
    print("=" * 78)
    print("DRY-RUN — Ingestão do questionário (Matriz + LC)")
    print("=" * 78)
    print(f"\nFicheiros: {len(files)}")
    print(f"Respostas ignoradas: {dict(skipped)}")
    print(f"Serviços distintos (entidade, nome normalizado): {len(display)}")
    by_ent = Counter(e for (e, _) in display)
    print(f"  por entidade: {dict(by_ent)}")

    print("\n── Agregação por indicador ──")
    per_ind = defaultdict(lambda: {"groups": 0, "codes": [], "cats": Counter(), "kind": None})
    for (ent, sn, y, m, chan, etl, kind), d in agg.items():
        pi = per_ind[etl]; pi["kind"] = kind; pi["groups"] += 1
        pi["codes"].extend(d["codes"]); pi["cats"].update(d["cats"])
    for etl in sorted(per_ind):
        pi = per_ind[etl]; codes = pi["codes"]
        if pi["kind"] in ("sim_nao", "agendamento", "nps"):
            print(f"  {etl:<20} [{pi['kind']:<11}] grupos={pi['groups']:<4} counts={dict(pi['cats']) or '-'} n={len(codes)}")
        else:
            mean = round(sum(codes) / len(codes), 2) if codes else None
            print(f"  {etl:<20} [{pi['kind']:<11}] grupos={pi['groups']:<4} n={len(codes):<5} média={mean}")

    print("\n── RÓTULOS NÃO MAPEADOS ──")
    print("  (nenhum)" if not unmapped else "  " + str({k: dict(v) for k, v in unmapped.items()}))

    print("\n── Distrito de Residência (2ª agregação, sem canal) ──")
    print(f"  respostas com distrito: {sum(geo_pop.values())} | sem distrito: {skipped.get('distrito_em_branco', 0)}")
    for distrito, n in geo_pop.most_common():
        print(f"  {distrito:<28} n={n}")

    print(f"\n── Linhas de measurement a gerar: {len(agg)} por canal + {len(agg_geo)} por distrito = {len(agg) + len(agg_geo)} ──")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--report", action="store_true")
    ap.add_argument("--sql")
    args = ap.parse_args()
    if args.sql:
        emit_sql(args.sql)
    else:
        report()


if __name__ == "__main__":
    main()
