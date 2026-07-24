#!/usr/bin/env python3
"""
Ingestão de dados operacionais reais da CML (Câmara Municipal de Lisboa) → org_cml.measurements.

Fontes (todas em docs/new/), cobrindo 3 dos 5 serviços da CML:

  1. Licenca_Utilizacao_CRM - 01062025 a 30062026.xlsx
     → Certidão de Licença de Utilização: volume mensal de pedidos (campo "Criado Em").
     NOTA: o ficheiro não tem coluna de canal. É assumido que o volume total de tickets do
     CRM representa o indicador genérico "Número de atendimentos presenciais por serviço"
     (mesmo indicador partilhado por todas as entidades na plataforma) — por isso estas
     linhas são marcadas is_provisional=TRUE, pendente de confirmação de canal pela CML.

  2. CML-Atendimentos-Caraterização-<Mês>26 - Anónimo.xlsx (6 ficheiros, Jan-Jun 2026)
     → Pedido de Certificado de Registo de Cidadão da UE, filtrando CATEGORIA =
     "Registo Cidadão da União Europeia – certificado":
       - Número de atendimentos (contagem de ID_ATENDIMENTO distintos por mês)
       - Tempo médio de atendimento (HORA FIM - HORA INICIO), em segundos
       - Tempo médio de espera (HORA INICIO - HORA DA SENHA), em segundos
     Dados completos e sem ambiguidade — NÃO marcados como provisórios.
     ATUALIZADO na migration 055 (2026-07-24, manual, fora deste script): pedido da cliente
     "Procura por serviço" → coluna SUBCATEGORIA (6 valores) dentro do mesmo filtro de
     CATEGORIA. Contagem por ID_ATENDIMENTO distinto, mesma metodologia acima. NOTA: ~400
     atendimentos têm mais do que 1 subcategoria registada, por isso a soma das
     subcategorias por mês fica um pouco acima do total mensal de "Número de
     atendimentos..." — is_provisional=TRUE nesse indicador novo por causa disto (o
     indicador de atendimentos em si continua NÃO provisório).

  3. CertificadosRegistoCidadaoUE_062025_062026.xlsx
     → Pedido de Certificado de Registo de Cidadão da UE: nº de certificados emitidos por mês.
     NOTA: o ficheiro só cobre Jan-Jun 2026 (o nome promete Jun2025-Jun2026 — meses em falta
     por confirmar com a CML) e tem 48 linhas exatamente duplicadas (removidas aqui) e 124
     pares certificado/processo repetidos em datas diferentes (mantidos como estão — podem
     ser reemissões legítimas, a confirmar). Por isso is_provisional=TRUE.
     ATUALIZADO na migration 055 (2026-07-24, manual, fora deste script): pedido da cliente
     "Number of requests by nationality" → coluna "Nacionalidade" (31 valores, mesma dedup de
     48 duplicados exatos). Novo indicador com category_counts por mês. NOTA de qualidade de
     dados: pelo menos 2 valores parecem truncados na fonte ("PAÍSES", "REPÚBLICA") —
     mantidos tal como estão, sem adivinhar o nome completo.

  4. Reclamacoes_Dashboard_*.xlsx (7 ficheiros, idênticos nas 3 pastas onde aparecem)
     → Pedido de Certificado de Registo de Cidadão da UE. SUBSTITUÍDO na migration 055
     (2026-07-24): o indicador "ue_reclamacoes" (f25b4776) que esta secção descrevia foi
     REMOVIDO do catálogo — ficou duplicado depois de a cliente pedir explicitamente os
     mesmos 7 indicadores standard de elogios/sugestões/reclamações já usados em Certidão e
     Feiras (migration 053). Os 4 registos que este indicador tinha (match exato de Assunto
     "Registo de cidadão da União Europeia - emissão de certificado", todos Reclamação +
     Resolvido) foram migrados para os indicadores standard "Número de reclamações
     recebidas"/"resolvidas". A função `load_reclamacoes_ue()` abaixo já não é chamada por
     `report()`/`emit_sql()` — fica só como referência histórica de como os 4 registos foram
     encontrados.

  5. Feiras_Representativo-01062025 a 30062026.xlsx
     → Solicitação de Lugar em Feira: nº de licenças atribuídas por feira e mês (campo
     "Data de Atribuição"). ATUALIZADO na migration 054 (2026-07-24): a coluna "Tipo
     deAtribuição" tem 3 valores (Ocasional Mensal=356, Extraordinária=260, Estudante=161,
     total 777) — o indicador "Número de licenças ocasionais atribuidas por feira e
     período" (853081cf) tinha sido ingerido aqui SEM filtrar por tipo (777, os 3 tipos
     juntos), o que estava errado dado o nome do indicador. Corrigido para conter só
     "Ocasional Mensal" (356). Criado o indicador "Número de licenças extraordinárias por
     feira" (260, só na Feira da Ladra) — não existia no catálogo. "Estudante" fica de fora
     (não pedido pela cliente). Dados completos e sem ambiguidade quanto ao valor em si —
     NÃO marcados como provisórios (a ambiguidade que existia era de categorização/catálogo,
     já resolvida).

DELIBERADAMENTE FORA DESTE SCRIPT (ver relatório de auditoria):
  - Pedido de Recolha de Monstros / Serviço de Recolha de Monstros: ATUALIZADO na migration
    056 (2026-07-24, manual, fora deste script). O ficheiro GOPI (LxM_Entradas/
    LxM_Resolvidas) continua sem NENHUM campo que distinga os 2 serviços catalogados na BD
    (tipologia é sempre "Remoção-Monstros-Pedido de recolha" nas duas sheets) — essa
    ambiguidade não foi resolvida. Mas a cliente pediu explicitamente 4 indicadores
    ("Número de pedidos"/"pedidos resolvidos" por canal e por freguesia) e David aprovou
    duplicar os mesmos dados pelos 2 serviços, is_provisional=TRUE — mesma convenção já
    usada para os dados UX destes 2 serviços (scripts/ingest_cml_ux.py). Os totais NÃO
    devem ser somados entre os 2 serviços (dupla contagem).
  - Indicadores de chamadas/IVR/canal de suporte: sem nenhuma fonte de dados nos ficheiros
    recebidos, para nenhum dos 4 serviços — ficam com o valor mock/fictício da migration 021.
  - Reclamações/elogios/sugestões e pedidos resolvidos/encaminhados para Certidão de Licença
    de Utilização: DEIXOU DE SER O CASO. A cliente pediu explicitamente estes 9 indicadores
    (2026-07-24) e foram aprovados e ingeridos — mas manualmente, via
    migrations/053_add_cml_certidao_reclamacoes_indicators.sql, não por este script (ver
    memória cml-data-gap-reconciliation). Só 2 registos em 5.428 correspondem ao assunto
    exato do serviço, por isso elogios/sugestões ficaram sem measurements (0 correspondências)
    e reclamações têm só 2 pontos (is_provisional=TRUE, recall do match por texto
    desconhecido).
  - Reclamações/elogios/sugestões e licenças extraordinárias/ocasionais mensais por feira
    para Solicitação de Lugar em Feira: idem, DEIXOU DE SER O CASO — pedido pela cliente
    (2026-07-24), aprovado e ingerido manualmente via
    migrations/054_add_cml_feiras_indicators_and_fix_ocasional.sql. Filtro por "Assunto
    (Novo)" ∈ {"Feiras", "Feira da Ladra"}: só 5 registos em 5.428 (0 elogios/sugestões, 5
    reclamações), is_provisional=TRUE pela mesma razão de sempre.
  - Nacionalidade dos certificados UE e subcategoria dos atendimentos UE: idem, pedido pela
    cliente (2026-07-24), aprovado e ingerido manualmente via
    migrations/055_add_cml_ue_indicators_and_remove_duplicate.sql (ver notas nas secções 2 e
    3 acima). Esta mesma migration também removeu o indicador "ue_reclamacoes" (f25b4776,
    ver secção 4 acima), substituído pelos indicadores standard de reclamações.

Uso:
    python3 scripts/ingest_cml_operational.py --report        # dry-run: valida e imprime
    python3 scripts/ingest_cml_operational.py --sql out.sql   # gera SQL de ingestão

NADA é escrito na base de dados por este script.
"""
import argparse, glob, json, os, unicodedata
from collections import Counter, defaultdict
from datetime import datetime

import openpyxl


def list_files(folder, prefix, suffix=".xlsx"):
    """glob() com padrões acentuados falha em macOS/APFS (NFC vs NFD) — comparar normalizado."""
    prefix_nfc = unicodedata.normalize("NFC", prefix)
    return sorted(
        os.path.join(folder, name)
        for name in os.listdir(folder)
        if unicodedata.normalize("NFC", name).startswith(prefix_nfc) and name.endswith(suffix)
    )

BASE = "docs/CML"  # pasta renomeada de docs/new; caminho corrigido em 2026-07-24

SCHEMA = "org_cml"

SERVICE_ID = {
    "certidao": "160b52e3-02ab-45d9-8564-24892cd71124",
    "ue": "fd42da38-0851-46d4-8afc-4c96cabbf7cb",
    "feira": "8e4b22cb-3602-4174-8b51-a19d1b3fcd09",
}

INDICATOR_ID = {
    "atendimentos_presenciais": "0b30f2bd-643e-409a-861b-61ad13913702",
    "ue_atendimentos": "a4705399-bc3e-4b48-9796-45a52aaccb4f",
    "ue_tma": "2df322e5-bcca-4c13-872c-33cd3f95a52c",
    "ue_tempo_espera": "c264285f-4076-4a35-9467-82642e2735b9",
    "ue_certificados_emitidos": "465031c9-0f52-4dc0-90dc-6dcd93b0660f",
    "feira_licencas": "853081cf-47f8-4d67-94d3-b600a302f3c3",
    # "ue_reclamacoes" (f25b4776) removido do catálogo na migration 055 — ver secção 4 do
    # cabeçalho. load_reclamacoes_ue() abaixo já não é chamada; fica só como referência.
}


def norm(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return " ".join(s.split())


# ── 1. Certidão de Licença de Utilização — volume mensal (CRM) ──────────────
def load_certidao_crm():
    path = f"{BASE}/Dados Matriz _ Certidão de Licença de Utilização - Envio de Ficheiros/Licenca_Utilizacao_CRM - 01062025 a 30062026.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Vista Localização Avançada ..."]
    months = Counter()
    for row in ws.iter_rows(min_row=2, values_only=True):
        criado = row[5]  # Criado Em
        if isinstance(criado, datetime):
            months[(criado.year, criado.month)] += 1
    wb.close()
    rows = []
    for (y, m), n in sorted(months.items()):
        rows.append({
            "service": SERVICE_ID["certidao"], "indicator": INDICATOR_ID["atendimentos_presenciais"],
            "year": y, "month": m, "value": float(n), "provisional": True,
            "source": "licenca_utilizacao_crm_2025_2026",
        })
    return rows, months


# ── 2. Atendimentos UE (CML-Atendimentos-Caraterização, 6 meses) ────────────
CATEGORIA_UE = "registo cidadao da uniao europeia – certificado"

def load_atendimentos_ue():
    folder = f"{BASE}/Dados Matriz _ Pedido de Certificado de Registo de Cidadão da União Europeia - Envio de Ficheiros"
    files = list_files(folder, "CML-Atendimentos-Caraterização-")
    seen_atend = defaultdict(set)          # (y,m) -> set(ID_ATENDIMENTO)
    dur_by_month = defaultdict(list)       # (y,m) -> [segundos atendimento]
    wait_by_month = defaultdict(list)      # (y,m) -> [segundos espera]
    for f in files:
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        ws = wb["Relatório"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            id_atend, hora_senha, hora_inicio, hora_fim, categoria = row[5], row[6], row[7], row[8], row[9]
            if norm(categoria) != norm(CATEGORIA_UE):
                continue
            if not isinstance(hora_inicio, datetime):
                continue
            y, m = hora_inicio.year, hora_inicio.month
            if id_atend not in seen_atend[(y, m)]:
                seen_atend[(y, m)].add(id_atend)
            if isinstance(hora_fim, datetime):
                dur_by_month[(y, m)].append((hora_fim - hora_inicio).total_seconds())
            if isinstance(hora_senha, datetime):
                wait_by_month[(y, m)].append((hora_inicio - hora_senha).total_seconds())
        wb.close()

    rows = []
    for (y, m), ids in sorted(seen_atend.items()):
        rows.append({
            "service": SERVICE_ID["ue"], "indicator": INDICATOR_ID["ue_atendimentos"],
            "year": y, "month": m, "value": float(len(ids)), "provisional": False,
            "source": "cml_atendimentos_caraterizacao_2026",
        })
    for (y, m), durs in sorted(dur_by_month.items()):
        if durs:
            rows.append({
                "service": SERVICE_ID["ue"], "indicator": INDICATOR_ID["ue_tma"],
                "year": y, "month": m, "value": round(sum(durs) / len(durs), 1), "provisional": False,
                "source": "cml_atendimentos_caraterizacao_2026",
            })
    for (y, m), waits in sorted(wait_by_month.items()):
        if waits:
            rows.append({
                "service": SERVICE_ID["ue"], "indicator": INDICATOR_ID["ue_tempo_espera"],
                "year": y, "month": m, "value": round(sum(waits) / len(waits), 1), "provisional": False,
                "source": "cml_atendimentos_caraterizacao_2026",
            })
    return rows, seen_atend


# ── 3. Certificados UE emitidos por mês ─────────────────────────────────────
def load_certificados_ue():
    path = f"{BASE}/Dados Matriz _ Pedido de Certificado de Registo de Cidadão da União Europeia - Envio de Ficheiros/CertificadosRegistoCidadaoUE_062025_062026.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Sheet1"]
    seen = set()
    months = Counter()
    dup_exact = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        data_str, cert, proc, nac = row[0], row[1], row[2], row[3]
        key = (data_str, cert, proc, nac)
        if key in seen:
            dup_exact += 1
            continue
        seen.add(key)
        d, m, y = data_str.split("/")
        months[(int(y), int(m))] += 1
    wb.close()
    rows = []
    for (y, m), n in sorted(months.items()):
        rows.append({
            "service": SERVICE_ID["ue"], "indicator": INDICATOR_ID["ue_certificados_emitidos"],
            "year": y, "month": m, "value": float(n), "provisional": True,
            "source": "certificados_registo_cidadao_ue_2026",
        })
    return rows, dup_exact, months


# ── 4. Reclamações UE (Reclamacoes_Dashboard, match exato de Assunto) ───────
# Histórico apenas — ver nota da secção 4 no cabeçalho do módulo. Este match já não
# alimenta nenhum indicador deste script (os 4 registos foram migrados manualmente para os
# indicadores standard de reclamações via migrations/055). Não devolve `rows` de
# measurement porque o indicador que usava (f25b4776) foi removido do catálogo.
TARGET_ASSUNTO_UE = "registo de cidadao da uniao europeia - emissao de certificado"

def load_reclamacoes_ue():
    folder = f"{BASE}/Dados Matriz _ Certidão de Licença de Utilização - Envio de Ficheiros"
    files = list_files(folder, "Reclamacoes_Dashboard_")
    matched_ids = []
    for f in files:
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        sn = "Dashboard_SER" if "Dashboard_SER" in wb.sheetnames else "SER_Dashboard"
        ws = wb[sn]
        rows_iter = ws.iter_rows(values_only=True)
        header = list(next(rows_iter))
        idx = {h: i for i, h in enumerate(header)}
        for row in rows_iter:
            if all(v is None for v in row):
                continue
            assunto = row[idx["Assunto (Novo)"]]
            if norm(assunto) != norm(TARGET_ASSUNTO_UE):
                continue
            if isinstance(row[idx["Criado Em"]], datetime):
                matched_ids.append(row[idx["Número de Pedido"]])
        wb.close()
    return matched_ids


# ── 5. Feiras — licenças atribuídas por feira e mês ─────────────────────────
def load_feiras():
    path = f"{BASE}/Dados Matriz _ Solicitação de Lugar em Feira - Envio de Ficheiros/Feiras_Representativo-01062025 a 30062026.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["listagem_de_ocupação"]
    counts = Counter()   # (feira, y, m) -> n
    for row in ws.iter_rows(min_row=2, values_only=True):
        feira, data_atrib = row[1], row[29]
        if isinstance(data_atrib, datetime):
            counts[(feira, data_atrib.year, data_atrib.month)] += 1
    wb.close()
    rows = []
    # Total agregado por mês (geo_level=NULL) — é esta série, não a repartição por feira,
    # que alimenta o valor "atual" mostrado no card do indicador (a UI escolhe a linha
    # com channel/geo_level NULL; a repartição por feira fica disponível para uma eventual
    # vista de detalhe, mas não substitui o agregado).
    total_by_month = Counter()
    for (feira, y, m), n in counts.items():
        total_by_month[(y, m)] += n
    for (y, m), n in sorted(total_by_month.items()):
        rows.append({
            "service": SERVICE_ID["feira"], "indicator": INDICATOR_ID["feira_licencas"],
            "year": y, "month": m, "value": float(n), "provisional": False,
            "source": "feiras_representativo_2025_2026",
        })
    for (feira, y, m), n in sorted(counts.items()):
        rows.append({
            "service": SERVICE_ID["feira"], "indicator": INDICATOR_ID["feira_licencas"],
            "year": y, "month": m, "geo_level": "feira", "geo_name": feira,
            "value": float(n), "provisional": False,
            "source": "feiras_representativo_2025_2026",
        })
    return rows, counts


def q(s):
    return "'" + str(s).replace("'", "''") + "'"


def report():
    print("=" * 78)
    print("DRY-RUN — Ingestão de dados operacionais reais da CML")
    print("=" * 78)

    r1, m1 = load_certidao_crm()
    print(f"\n[1] Certidão CRM — {len(r1)} meses, {sum(m1.values())} pedidos totais")
    for (y, m), n in sorted(m1.items()):
        print(f"    {y}-{m:02d}: {n}")

    r2, atend2 = load_atendimentos_ue()
    total_atend = sum(len(v) for v in atend2.values())
    print(f"\n[2] Atendimentos UE (CML-Atendimentos) — {total_atend} atendimentos distintos, {len(r2)} linhas de measurement")
    for (y, m), ids in sorted(atend2.items()):
        print(f"    {y}-{m:02d}: {len(ids)} atendimentos")

    r3, dup3, m3 = load_certificados_ue()
    print(f"\n[3] Certificados UE emitidos — {sum(m3.values())} certificados únicos ({dup3} duplicados exatos removidos)")
    for (y, m), n in sorted(m3.items()):
        print(f"    {y}-{m:02d}: {n}")

    ids4 = load_reclamacoes_ue()
    print(f"\n[4] Reclamações UE (match exato de Assunto, histórico) — {len(ids4)} registos: {ids4}")
    print("    Já não gera measurement por este script — ver migrations/055 e nota da secção 4 no cabeçalho.")

    r5, c5 = load_feiras()
    total_feiras = sum(c5.values())
    print(f"\n[5] Feiras — {total_feiras} licenças atribuídas em {len(c5)} combinações (feira, mês)")
    by_feira = Counter()
    for (feira, y, m), n in c5.items():
        by_feira[feira] += n
    for feira, n in by_feira.most_common():
        print(f"    {feira}: {n}")

    all_rows = r1 + r2 + r3 + r5
    print(f"\n── Total de linhas de measurement a gerar: {len(all_rows)} ──")
    print("── Deliberadamente fora deste script (ver cabeçalho para detalhe): Monstros (Pedido/Serviço), chamadas/IVR/canal de suporte. Reclamações/elogios/sugestões, nacionalidade e subcategoria da UE, e os indicadores novos de Certidão/Feiras foram ingeridos manualmente via migrations 053-055. ──")


def add_snapshot_rows(all_rows):
    """A UI mostra o valor 'atual' do card escolhendo a linha com channel/geo_level NULL,
    sem ordenar por período — por isso, além do histórico mensal, é preciso 1 linha extra
    por (serviço, indicador, geo_level, geo_name) com month=NULL, igual ao mês mais recente,
    que substitui (via ON CONFLICT, mesma chave da migration 021) o valor mock antigo."""
    latest = {}
    for row in all_rows:
        key = (row["service"], row["indicator"], row.get("geo_level"), row.get("geo_name"))
        ym = (row["year"], row["month"])
        if key not in latest or ym > latest[key]["_ym"]:
            snap = dict(row)
            snap["_ym"] = ym
            latest[key] = snap
    snapshots = []
    for snap in latest.values():
        snap = dict(snap)
        snap.pop("_ym")
        snap["month"] = None
        snapshots.append(snap)
    return all_rows + snapshots


def emit_sql(out):
    r1, _ = load_certidao_crm()
    r2, _ = load_atendimentos_ue()
    r3, _, _ = load_certificados_ue()
    r5, _ = load_feiras()
    all_rows = add_snapshot_rows(r1 + r2 + r3 + r5)

    L = []
    L.append("-- Gerado por scripts/ingest_cml_operational.py — NÃO editar à mão.")
    L.append("-- Fontes: docs/new/ (ver cabeçalho do script para detalhe por indicador).")
    L.append("-- Substitui os valores mock da migration 021 nos indicadores/serviços/meses cobertos por dados reais.")
    L.append("-- Inclui, além do histórico mensal, uma linha month=NULL por série = valor do mês mais recente\n"
              "-- (é essa linha que a UI usa como valor 'atual' do card — ver src/app/indicadores/[id]/page.tsx).\n")

    values = []
    del_keys = []
    for row in all_rows:
        geo_level = row.get("geo_level")
        geo_name = row.get("geo_name")
        month = "NULL" if row["month"] is None else row["month"]
        values.append(
            f"({q(row['service'])}::uuid,{q(row['indicator'])}::uuid,{row['year']},{month},"
            f"{'NULL' if geo_level is None else q(geo_level)},{'NULL' if geo_name is None else q(geo_name)},"
            f"{row['value']!r},{str(row['provisional']).upper()},{q(row['source'])})"
        )
        del_keys.append((row['service'], row['indicator'], row['year'], month, geo_level, geo_name))

    # ON CONFLICT não é fiável para as linhas snapshot (month=NULL): o Postgres trata NULL
    # como distinto de NULL num índice único, logo duas linhas com month=NULL nunca colidem
    # entre si, mesmo com o resto da chave igual — por isso apagamos primeiro (incl. o mock
    # antigo da migration 021) usando IS NOT DISTINCT FROM, e só depois inserimos.
    L.append(
        "DELETE FROM org_cml.measurements m\n"
        "USING (VALUES\n  " + ",\n  ".join(
            f"({q(s)}::uuid,{q(i)}::uuid,{y},{m},{'NULL' if gl is None else q(gl)},{'NULL' if gn is None else q(gn)})"
            for s, i, y, m, gl, gn in del_keys
        ) + "\n) AS v(service_id, indicator_id, year, month, geo_level, geo_name)\n"
        "WHERE m.service_id = v.service_id AND m.indicator_id = v.indicator_id AND m.year = v.year\n"
        "  AND m.month IS NOT DISTINCT FROM v.month AND m.channel IS NULL\n"
        "  AND m.geo_level IS NOT DISTINCT FROM v.geo_level AND m.geo_name IS NOT DISTINCT FROM v.geo_name;\n"
    )
    L.append(
        "INSERT INTO org_cml.measurements "
        "(service_id, indicator_id, year, month, geo_level, geo_name, value, is_provisional, source_file)\n"
        "VALUES\n  " + ",\n  ".join(values) + ";"
    )

    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(L) + "\n")
    print(f"SQL escrito em {out}: {len(values)} linhas (histórico mensal + snapshots month=NULL).")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--report", action="store_true")
    ap.add_argument("--sql")
    args = ap.parse_args()
    if args.sql:
        emit_sql(args.sql)
    else:
        report()


if __name__ == "__main__":
    main()
