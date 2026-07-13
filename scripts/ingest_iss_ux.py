#!/usr/bin/env python3
"""
Ingestão do questionário de experiência do utilizador multi-canal da ISS → measurements.

Ficheiro fonte: docs/ISS - UX experience (multi-channel).xlsx (30 respostas individuais,
4 serviços). Layout de colunas DIFERENTE do questionário Matriz/LC anterior — por isso
este script tem o seu próprio mapeamento de colunas, calibrado a este ficheiro.

Diferença crítica em relação a scripts/ingest_survey.py: este script NUNCA apaga
medições existentes da ISS — só insere as novas agregações (a ISS já tem dados reais
de outras fontes que têm de ficar intactos).

Uso:
    python3 scripts/ingest_iss_ux.py --report        # dry-run: só valida e imprime
    python3 scripts/ingest_iss_ux.py --sql out.sql   # gera SQL de ingestão (INSERT only)

Qualquer rótulo de resposta não previsto é reportado como UNMAPPED — nunca adivinhado.
"""
import argparse, json, sys, unicodedata
from collections import defaultdict, Counter
from datetime import datetime

import openpyxl

FILE_PATH = "docs/ISS - UX experience (multi-channel).xlsx"

# ── Normalização de texto ────────────────────────────────────────────────────
def norm(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return " ".join(s.split())

# ── Escalas texto → código 1-5 (idênticas ao ingest_survey.py — mesmo instrumento) ──
FACILIDADE = {"muito dificil": 1, "dificil": 2, "nem facil nem dificil": 3, "neutro": 3, "facil": 4, "muito facil": 5}
RAPIDEZ = {"muito lento": 1, "lento": 2, "nem lento nem rapido": 3, "rapido": 4, "muito rapido": 5}
CONCORDANCIA = {"discordo totalmente": 1, "discordo parcialmente": 2, "nao concordo nem discordo": 3, "concordo parcialmente": 4, "concordo totalmente": 5}
CONCORDANCIA_DUR = {"discordo totalmente": 1, "discordo": 2, "nao concordo nem discordo": 3, "concordo": 4, "concordo totalmente": 5}
QUALIDADE = {"muito mau": 1, "mau": 2, "nem mau nem bom": 3, "bom": 4, "muito bom": 5}
SATISFACAO = {"muito insatisfatorio": 1, "insatisfatorio": 2, "nem insatisfatorio nem satisfatorio": 3, "satisfatorio": 4, "muito satisfatorio": 5}
ESPERA = {"muito demorado/inaceitavel": 1, "demorado": 2, "tempo aceitavel": 3, "rapido": 4, "muito rapido": 5}
EXPECTATIVAS = {
    "nao correspondeu nada as expectativas": 1,
    "nao correspondeu em parte nem totalmente as expectativas": 2,
    "correspondeu em parte as expectativas": 3,
    "correspondeu totalmente as expectativas": 4,
    "superou as expectativas": 5,
}
EXCLUDE = {"nao utilizei este canal", "nao se aplica", "nao se aplica.", "prefiro nao responder", "ainda aguardo conclusao/resposta"}

# ── Mapa coluna Excel (1-indexed) → indicador (etl_column_key) ──────────────
# kind: likert | scale10 | nps | sim_nao | agendamento
COLMAP = [
    (9,  "ux_resolved",        "sim_nao",     None,             None),
    (11, "ux_speed",           "likert",      RAPIDEZ,          None),
    (12, "ux_channel_ease",    "likert",      FACILIDADE,       "Presencial"),
    (13, "ux_channel_ease",    "likert",      FACILIDADE,       "Telefone"),
    (14, "ux_channel_ease",    "likert",      FACILIDADE,       "Digital/Online"),   # SSD/Website
    (15, "ux_channel_ease",    "likert",      FACILIDADE,       "App"),
    (16, "ux_channel_ease",    "likert",      FACILIDADE,       "Videochamada"),
    (17, "ux_channel_ease",    "likert",      FACILIDADE,       "Chatbox"),          # Chatbot/Assistente virtual
    (18, "ux_channel_ease",    "likert",      FACILIDADE,       "E-clic"),           # canal próprio da ISS
    (19, "ux_channel_ease",    "likert",      FACILIDADE,       "Outro"),
    (20, "ux_findability",     "likert",      FACILIDADE,       None),
    (21, "ux_language_simple", "likert",      CONCORDANCIA,     None),
    (22, "ux_language_clear",  "likert",      CONCORDANCIA,     None),
    (23, "ux_response_clarity","sim_nao",     None,             None),
    (27, "ux_ease",            "likert",      FACILIDADE,       None),
    (28, "ux_scheduling",      "agendamento", None,             None),
    (29, "ux_scheduling_ease", "likert",      FACILIDADE,       None),
    (30, "ux_scheduling_date", "likert",      FACILIDADE,       None),
    (33, "ux_wait_time",       "likert",      ESPERA,           None),
    (34, "ux_duration",        "likert",      CONCORDANCIA_DUR, None),
    (35, "ux_clarity_info",    "likert",      QUALIDADE,        None),
    (36, "ux_knowledge",       "likert",      QUALIDADE,        None),
    (37, "ux_routing",         "likert",      QUALIDADE,        None),
    (38, "ux_fairness",        "likert",      CONCORDANCIA,     None),
    (39, "ux_courtesy",        "likert",      SATISFACAO,       None),
    (40, "ux_csat",            "scale10",     None,             None),
    (41, "ux_expectations",    "likert",      EXPECTATIVAS,     None),
    (42, "ux_nps",             "nps",         None,             None),
]
# Colunas intencionalmente NÃO ingeridas (sem indicador correspondente no catálogo,
# ou texto aberto / metadados demográficos):
#   8  Que canais utilizou (resumo multi-select, não é pergunta de escala)
#   10 Se respondeu "Não"... porquê (texto aberto)
#   24-26 Compreender a decisão / estado do pedido / próximos passos (sub-dimensões
#         de transparência sem indicador dedicado — mesmo critério do ingest_survey.py)
#   31 Caso não tenha conseguido agendar... que problemas (texto aberto)
#   32 Foi atendido por um operador humano (SEM indicador correspondente no catálogo)
#   43 Sugestão para melhorar o serviço (texto aberto)
#   44-48 Identificação / primeira vez / idade / profissão / escolaridade (demografia)
NOTE_SKIPPED_COLS = {
    8: "Que canais utilizou (resumo multi-select, sem indicador dedicado)",
    10: "Se respondeu Não... porquê (texto aberto)",
    24: "Compreender a decisão (sub-dimensão sem indicador dedicado)",
    25: "Saber o estado do pedido (sub-dimensão sem indicador dedicado)",
    26: "Compreender os próximos passos (sub-dimensão sem indicador dedicado)",
    31: "Problemas no agendamento (texto aberto)",
    32: "Foi atendido por um operador humano (SEM indicador correspondente no catálogo)",
    43: "Sugestão para melhorar o serviço (texto aberto)",
}

DISTRITO_COL = 49
SERVICE_COL = 7
COMPLETION_COL = 3

SIM = "Sim"; NAO = "Não"; NA = "Não aplicável"; NMT = "Não, mas tentei"

ORG_ID_ISS = "dca9b165-c7d9-44b7-afd8-38ebc13e604c"
SCHEMA_ISS = "org_iss"

# ── Serviços do ficheiro → serviço já existente na BD (decidido com o David) ─
# Os nomes no ficheiro não coincidem carácter-a-carácter com os da BD; mapeados
# explicitamente para não depender de correspondência automática por nome.
SERVICE_OVERRIDE = {
    "pensao de velhice": ("cf1becbf-0157-497e-bda8-62fce6645f9e", "Pensão de velhice"),
    "pensao de invalidez": ("e5ff21c8-4f06-4eab-9b81-489139026e13", "Pensão de invalidez"),
    "abono de familia para criancas e jovens": ("bcafb8a1-a3db-48c1-a6bb-e48ad6375b81", "Abono de Família para Crianças e Jovens - Majorações"),
    "subsidio parental inicial": ("cc964c9c-47ff-4b17-867e-4fb09448c741", "Abono parental inicial"),
}


def period_of(row):
    v = row[COMPLETION_COL - 1]
    if isinstance(v, datetime):
        return v.year, v.month
    return None, None


def load():
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    ws = wb["Sheet1"]

    agg = defaultdict(lambda: {"codes": [], "cats": Counter()})       # key=(sn, y, m, chan, etl, kind)
    agg_geo = defaultdict(lambda: {"codes": [], "cats": Counter()})   # key=(sn, y, m, geo_name, etl, kind)
    pop = Counter()                       # (sn, y, m) -> nº respostas
    geo_pop = Counter()
    display = defaultdict(Counter)        # sn -> Counter(nome cru)
    unmapped_service = Counter()
    unmapped = defaultdict(Counter)
    skipped = Counter()

    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(v in (None, "") for v in row):
            continue

        raw_serv = row[SERVICE_COL - 1]
        if not raw_serv:
            skipped["servico_em_branco"] += 1
            continue
        sn = norm(str(raw_serv))
        if sn not in SERVICE_OVERRIDE:
            unmapped_service[str(raw_serv).strip()] += 1
            continue

        y, m = period_of(row)
        if not y:
            skipped["periodo_invalido"] += 1
            continue

        pop[(sn, y, m)] += 1
        display[sn][str(raw_serv).strip()] += 1

        distrito_raw = row[DISTRITO_COL - 1]
        geo_name = str(distrito_raw).strip() if distrito_raw not in (None, "") else None
        if geo_name:
            geo_pop[geo_name] += 1
        else:
            skipped["distrito_em_branco"] += 1

        for col, etl, kind, scale, chan in COLMAP:
            raw = row[col - 1]
            if raw in (None, ""):
                continue
            key = (sn, y, m, chan, etl, kind)
            geo_key = (sn, y, m, geo_name, etl, kind) if geo_name else None

            if kind in ("sim_nao", "agendamento"):
                nv = norm(raw)
                cat = None
                if nv == "sim":
                    cat = SIM
                elif nv == "nao" or nv == "nao recebi resposta":
                    cat = NAO
                elif nv == "nao aplicavel":
                    cat = NA
                elif nv in ("nao, mas tentei", "nao mas tentei"):
                    cat = NMT
                if cat is None:
                    unmapped[etl][str(raw).strip()] += 1
                else:
                    agg[key]["cats"][cat] += 1
                    if geo_key:
                        agg_geo[geo_key]["cats"][cat] += 1
            elif kind == "likert":
                nv = norm(raw)
                if nv in EXCLUDE:
                    continue
                code = scale.get(nv)
                if code is None:
                    unmapped[etl][str(raw).strip()] += 1
                else:
                    agg[key]["codes"].append(code)
                    if geo_key:
                        agg_geo[geo_key]["codes"].append(code)
            elif kind in ("scale10", "nps"):
                try:
                    fv = float(raw)
                except Exception:
                    unmapped[etl][str(raw).strip()] += 1
                else:
                    agg[key]["codes"].append(fv)
                    if geo_key:
                        agg_geo[geo_key]["codes"].append(fv)

    return dict(agg), dict(agg_geo), pop, geo_pop, display, unmapped, unmapped_service, skipped


def compute(kind, d):
    codes, cats = d["codes"], d["cats"]
    if kind in ("likert", "scale10"):
        n = len(codes)
        return (round(sum(codes) / n, 2) if n else None), None, n
    if kind == "nps":
        n = len(codes)
        if not n:
            return None, None, 0
        promo = sum(1 for v in codes if v >= 9)
        detr = sum(1 for v in codes if v <= 6)
        pas = n - promo - detr
        nps = round(100 * (promo - detr) / n, 1)
        return nps, {"Promotores": promo, "Passivos": pas, "Detratores": detr}, n
    if kind == "sim_nao":
        s, nn, na = cats[SIM], cats[NAO], cats[NA]
        den = s + nn
        val = round(100 * s / den, 1) if den else None
        cc = {SIM: s, NAO: nn}
        if na:
            cc[NA] = na
        return val, cc, s + nn + na
    if kind == "agendamento":
        s, nn, nmt = cats[SIM], cats[NAO], cats[NMT]
        return None, {SIM: s, NAO: nn, NMT: nmt}, s + nn + nmt
    return None, None, 0


def q(s):
    return "'" + str(s).replace("'", "''") + "'"


def report():
    agg, agg_geo, pop, geo_pop, display, unmapped, unmapped_service, skipped = load()
    print("=" * 78)
    print("DRY-RUN — Ingestão UX multi-canal da ISS")
    print("=" * 78)
    print(f"\nRespostas ignoradas: {dict(skipped)}")
    print(f"Serviços mapeados: {len(display)} / {len(SERVICE_OVERRIDE)} esperados")
    for sn, c in display.items():
        _, canon = SERVICE_OVERRIDE[sn]
        print(f"  {c.most_common(1)[0][0]!r} → {canon!r}  (n={sum(c.values())})")

    print("\n── Colunas intencionalmente não ingeridas ──")
    for col, why in NOTE_SKIPPED_COLS.items():
        print(f"  col {col}: {why}")

    print("\n── Serviços NÃO reconhecidos (fora do mapeamento acordado) ──")
    print("  (nenhum)" if not unmapped_service else "  " + str(dict(unmapped_service)))

    print("\n── Agregação por indicador ──")
    per_ind = defaultdict(lambda: {"groups": 0, "codes": [], "cats": Counter(), "kind": None})
    for (sn, y, m, chan, etl, kind), d in agg.items():
        pi = per_ind[etl]; pi["kind"] = kind; pi["groups"] += 1
        pi["codes"].extend(d["codes"]); pi["cats"].update(d["cats"])
    for etl in sorted(per_ind):
        pi = per_ind[etl]; codes = pi["codes"]
        if pi["kind"] in ("sim_nao", "agendamento", "nps"):
            print(f"  {etl:<20} [{pi['kind']:<11}] grupos={pi['groups']:<4} counts={dict(pi['cats']) or '-'} n={len(codes)}")
        else:
            mean = round(sum(codes) / len(codes), 2) if codes else None
            print(f"  {etl:<20} [{pi['kind']:<11}] grupos={pi['groups']:<4} n={len(codes):<5} média={mean}")

    print("\n── RÓTULOS NÃO MAPEADOS (respostas com texto inesperado) ──")
    print("  (nenhum)" if not unmapped else "  " + str({k: dict(v) for k, v in unmapped.items()}))

    print("\n── Distrito de Residência (2ª agregação, sem canal) ──")
    print(f"  respostas com distrito: {sum(geo_pop.values())} | sem distrito: {skipped.get('distrito_em_branco', 0)}")
    for distrito, n in geo_pop.most_common():
        print(f"  {distrito:<28} n={n}")

    print(f"\n── Linhas de measurement a gerar: {len(agg)} por canal + {len(agg_geo)} por distrito = {len(agg) + len(agg_geo)} ──")
    print("\n── Nenhuma medição existente da ISS é apagada por este script (só INSERT). ──")


def emit_sql(out):
    agg, agg_geo, pop, geo_pop, display, unmapped, unmapped_service, skipped = load()
    assert not unmapped, f"Há rótulos não mapeados, abortar: {dict(unmapped)}"
    assert not unmapped_service, f"Há serviços não reconhecidos, abortar: {dict(unmapped_service)}"

    L = []
    L.append("-- Gerado por scripts/ingest_iss_ux.py — NÃO editar à mão.")
    L.append(f"-- Fonte: {FILE_PATH}")
    L.append("-- Só insere medições novas; não apaga nenhuma medição existente da ISS.\n")

    rows = []
    for (sn, y, m, chan, etl, kind), d in sorted(agg.items(), key=lambda kv: tuple(str(x) for x in kv[0])):
        value, cc, resp = compute(kind, d)
        service_id, _ = SERVICE_OVERRIDE[sn]
        inq = pop[(sn, y, m)]
        vstr = "NULL" if value is None else repr(float(value))
        cstr = "NULL" if cc is None else q(json.dumps(cc, ensure_ascii=False))
        chstr = "NULL" if chan is None else q(chan)
        rows.append(f"({q(service_id)}::uuid,{q(etl)},{y},{m},{chstr},NULL,NULL,{vstr},{cstr},{resp},{inq})")
    for (sn, y, m, geo_name, etl, kind), d in sorted(agg_geo.items(), key=lambda kv: tuple(str(x) for x in kv[0])):
        value, cc, resp = compute(kind, d)
        service_id, _ = SERVICE_OVERRIDE[sn]
        inq = pop[(sn, y, m)]
        vstr = "NULL" if value is None else repr(float(value))
        cstr = "NULL" if cc is None else q(json.dumps(cc, ensure_ascii=False))
        rows.append(f"({q(service_id)}::uuid,{q(etl)},{y},{m},NULL,{q('distrito')},{q(geo_name)},{vstr},{cstr},{resp},{inq})")

    if rows:
        L.append(
            "INSERT INTO org_iss.measurements "
            "(service_id, indicator_id, year, month, channel, geo_level, geo_name, value, category_counts, "
            "total_respondentes, total_inquiridos, source_file)\nSELECT v.service_id, i.id, v.year, v.month, "
            "v.channel, v.geo_level, v.geo_name, v.value::numeric, v.cc::jsonb, v.resp, v.inq, 'iss_ux_multicanal_2026'\n"
            "FROM (VALUES\n  " + ",\n  ".join(rows) +
            "\n) AS v(service_id, etl, year, month, channel, geo_level, geo_name, value, cc, resp, inq)\n"
            "JOIN public.indicators i ON i.etl_column_key = v.etl\n"
            "ON CONFLICT (service_id, indicator_id, year, month, channel, geo_level, geo_name) DO NOTHING;"
        )

    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(L) + "\n")
    print(f"SQL escrito em {out}: {len(rows)} medições.")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--report", action="store_true")
    ap.add_argument("--sql")
    args = ap.parse_args()
    if args.sql:
        emit_sql(args.sql)
    else:
        report()


if __name__ == "__main__":
    main()
