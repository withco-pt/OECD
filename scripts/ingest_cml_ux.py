#!/usr/bin/env python3
"""
Ingestão dos 3 questionários UX novos da CML → org_cml.measurements.

Ficheiros (docs/new/Questionários UX CML/), um 3º layout de colunas, distinto tanto do
"Matriz+LC" (scripts/ingest_survey.py) como do "ISS multicanal" (scripts/ingest_iss_ux.py):

  A. "Questionário Presencial - Avaliação da experiência de utilização dos serviços.xlsx"
     5 respostas, cobre 3 serviços (coluna 6 identifica qual).
  B. "Questionário UX Online -Certidão de Licença de Utilização - ...xlsx"
     2 respostas, sem coluna de serviço — implícito no nome do ficheiro (Certidão).
  C. "Questuinário UX Recolha de Monstros - ...xlsx"
     8 respostas, sem coluna de serviço — ver nota sobre Monstros abaixo. Layout com 1
     canal extra ("Online - Na Minha Rua Lx") e cabeçalhos de algumas colunas corrompidos
     no export (vieram como números "2".."9" em vez do texto da pergunta) — confirmados
     por posição e comparação com os outros 2 ficheiros (mesma ordem de perguntas).

Amostra total: 15 respostas (5+2+8) — TODAS foram inspecionadas manualmente, não há mais
nenhuma resposta nestes ficheiros.

Só são ingeridos indicadores com etl_column_key (mesmo critério de scripts/ingest_survey.py
e scripts/ingest_iss_ux.py) — sub-dimensões sem indicador dedicado (compreender decisão/
estado/próximos passos, "assistente digital", "realizou atendimento") são reportadas como
skipped mas não ingeridas, para consistência com os scripts existentes.

Rótulos de escala: este layout usa formulações mais curtas que os dicionários já
codificados (vírgulas em vez de "nem X nem Y", sem sufixo "às expectativas", Concordo/
Discordo sem qualificador). Os dicionários abaixo são extensões explícitas dos já usados
em ingest_survey.py/ingest_iss_ux.py — cada entrada nova está comentada com o rótulo
literal encontrado nos ficheiros. Qualquer rótulo não previsto é reportado como UNMAPPED.

Serviços:
  - "Certidão de Licença de Utilização" (ficheiro A) e ficheiro B → mapeiam 1:1 ao nome
    já existente na BD.
  - "Licença Ocasional para Venda na Feira da Ladra, das Galinheiras e do Relógio"
    (ficheiro A) → mapeado a "Solicitação de Lugar em Feira".
  - "Certificado de Registo de Cidadão da União Europeia" (ficheiro A) → mapeado a
    "Pedido de Certificado de Registo de Cidadão da União Europeia".
  - Ficheiro C (Monstros): tal como no ficheiro operacional GOPI, NADA no questionário
    distingue "Pedido de Recolha de Monstros" de "Serviço de Recolha de Monstros" — as
    perguntas cobrem a experiência ponta-a-ponta (pedido E recolha). Por decisão already
    tomada no âmbito de "avançar com o máximo de dados possível": estas respostas são
    ingeridas em AMBOS os serviços, marcadas is_provisional=TRUE, com nota explícita de
    que a divisão Pedido/Serviço não está confirmada — pendente de esclarecimento CML/ARTE.

Uso:
    python3 scripts/ingest_cml_ux.py --report        # dry-run: valida e imprime
    python3 scripts/ingest_cml_ux.py --sql out.sql   # gera SQL de ingestão

NADA é escrito na base de dados por este script.
"""
import argparse, json, unicodedata
from collections import defaultdict, Counter
from datetime import datetime

import openpyxl

BASE = "docs/CML/Questionários UX CML"

SERVICE_ID = {
    "certidao": ("160b52e3-02ab-45d9-8564-24892cd71124", "Certidão de Licença de Utilização"),
    "ue": ("fd42da38-0851-46d4-8afc-4c96cabbf7cb", "Pedido de Certificado de Registo de Cidadão da União Europeia"),
    "feira": ("8e4b22cb-3602-4174-8b51-a19d1b3fcd09", "Solicitação de Lugar em Feira"),
    "monstros_pedido": ("636ec47d-d22c-48b3-9db2-a5c0c1af2cb5", "Pedido de Recolha de Monstros"),
    "monstros_servico": ("8908d4e4-cf22-4e29-ad98-55233f5b9e1a", "Serviço de Recolha de Monstros"),
}

# Nomes de serviço tal como aparecem no ficheiro A (coluna 6) → chave interna acima.
SERVICE_LABEL_MAP = {
    "certidao de licenca de utilizacao": "certidao",
    "certificado de registo de cidadao da uniao europeia": "ue",
    "licenca ocasional para venda na feira da ladra das galinheiras e do relogio": "feira",
}


def norm(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    s = s.replace(",", "")
    return " ".join(s.split())


# ── Dicionários de escala texto → código 1-5 ────────────────────────────────
# Base: iguais aos já usados em ingest_survey.py / ingest_iss_ux.py.
FACILIDADE = {"muito dificil": 1, "dificil": 2, "nem facil nem dificil": 3, "facil": 4, "muito facil": 5}
RAPIDEZ = {
    "muito lento": 1, "lento": 2, "nem lento nem rapido": 3,
    "nem rapido nem lento": 3,   # NOVO — "Nem rápido, nem lento" (ordem trocada face ao dicionário antigo)
    "rapido": 4, "muito rapido": 5,
}
QUALIDADE = {
    "muito mau": 1, "mau": 2, "nem mau nem bom": 3,
    "nem bom nem mau": 3,        # NOVO — "Nem bom, nem mau" (ordem trocada)
    "bom": 4, "muito bom": 5,
}
SATISFACAO = {"muito insatisfatorio": 1, "insatisfatorio": 2, "nem insatisfatorio nem satisfatorio": 3, "satisfatorio": 4, "muito satisfatorio": 5}
ESPERA = {"muito demorado/inaceitavel": 1, "demorado": 2, "tempo aceitavel": 3, "rapido": 4, "muito rapido": 5}
# CONCORDANCIA_CURTA — este questionário usa sempre "Concordo"/"Discordo" sem qualificador
# ("parcialmente"), ao contrário do Matriz+LC/ISS. Usada para linguagem clara/simples,
# justiça no atendimento E duração do atendimento (as 3 usam a mesma formulação aqui).
CONCORDANCIA_CURTA = {
    "discordo totalmente": 1, "discordo": 2,
    "discordo parcialmente": 2,   # NOVO — visto 1x em ux_fairness, tratado ao mesmo nível de "discordo"
    "nao concordo nem discordo": 3,
    "concordo": 4,
    "concordo parcialmente": 4,   # NOVO — visto 1x em ux_fairness, tratado ao mesmo nível de "concordo"
    "concordo totalmente": 5,
}
EXPECTATIVAS = {
    "nao correspondeu nada as expectativas": 1,
    "nao correspondeu nada": 1,                              # NOVO — sem sufixo
    "nao correspondeu em parte nem totalmente as expectativas": 2,
    "nao correspondeu em parte nem totalmente": 2,            # NOVO — sem sufixo
    "correspondeu em parte as expectativas": 3,
    "correspondeu totalmente as expectativas": 4,
    "correspondeu totalmente": 4,                             # NOVO — sem sufixo
    "superou as expectativas": 5,
}
EXCLUDE = {
    "nao utilizei este canal", "nao usei este canal",  # NOVO — "Não usei este canal"
    "nao se aplica", "nao se aplica.", "nao aplicavel",  # NOVO — "Não aplicável" (likert)
    "prefiro nao responder", "ainda aguardo conclusao/resposta",
}

SIM = "Sim"; NAO = "Não"; NA = "Não aplicável"; NMT = "Não, mas tentei"

CHANNEL_MAP = {
    "presencial": "Presencial",
    "online - lisboa.pt": "Digital/Online",
    "online - lisboapt": "Digital/Online",
    "online - chat": "Chatbox",
    "telefone": "Telefone",
    "online - na minha rua lx": "Na Minha Rua Lx",  # canal próprio do serviço de Monstros
}


def channel_of(header_label):
    return CHANNEL_MAP.get(norm(header_label), header_label)


# ── Ligação canal↔indicadores (2026-07-22, a pedido da cliente) ─────────────
# A pergunta "Que canais utilizou para realizar este serviço?" (checklist multi-select,
# nunca ingerida) permite atribuir as restantes ~20 perguntas (satisfação, resolução,
# etc. — hoje só agregadas, sem canal) ao canal do inquirido, quando ele usou UM SÓ
# canal. Para quem usou vários canais, não sabemos qual gerou a resposta — não se
# atribui (fica só na agregação, como já acontecia antes). Só se mapeiam rótulos
# EXATAMENTE observados nos ficheiros (nunca adivinhados); qualquer rótulo novo fica
# UNMAPPED e é reportado, sem abortar a ingestão.
CHECKLIST_CHANNEL_MAP = {
    "atendimento presencial": "Presencial",
    "online - lisboa․pt": "Online - Lisboa.pt",  # rótulo usa "․" (dot leader), não "."
    "telefone": "Telefone",
    "online - na minha rua lx": "Online - Na Minha Rua Lx",
}


def respondent_channel(raw_checklist, channel_map, unmapped_counter):
    """Devolve o canal único do inquirido, ou None (vazio, multicanal, ou rótulo desconhecido)."""
    if raw_checklist in (None, ""):
        return None
    parts = [p.strip() for p in str(raw_checklist).split(";") if p.strip()]
    if not parts:
        return None
    canon = set()
    for p in parts:
        c = channel_map.get(norm(p))
        if c is None:
            unmapped_counter[p] += 1
            return None
        canon.add(c)
    return canon.pop() if len(canon) == 1 else None


# kind: likert | scale10 | nps | sim_nao | agendamento
# Cada entrada: (col 1-indexed, etl_column_key, kind, dicionário, canal|None)

COLMAP_A = [  # Questionário Presencial — 39 colunas, serviço na coluna 6
    (10, "ux_resolved", "sim_nao", None, None),
    (12, "ux_speed", "likert", RAPIDEZ, None),
    (13, "ux_channel_ease", "likert", FACILIDADE, "Presencial"),
    (14, "ux_channel_ease", "likert", FACILIDADE, "Online - Lisboa.pt"),
    (15, "ux_channel_ease", "likert", FACILIDADE, "Online - chat"),
    (16, "ux_channel_ease", "likert", FACILIDADE, "Telefone"),
    (17, "ux_findability", "likert", FACILIDADE, None),
    (18, "ux_language_clear", "likert", CONCORDANCIA_CURTA, None),
    (19, "ux_language_simple", "likert", CONCORDANCIA_CURTA, None),
    (20, "ux_response_clarity", "sim_nao", None, None),
    (24, "ux_ease", "likert", FACILIDADE, None),
    (25, "ux_scheduling", "agendamento", None, None),
    (26, "ux_scheduling_ease", "likert", FACILIDADE, None),
    (27, "ux_scheduling_date", "likert", FACILIDADE, None),
    (29, "ux_wait_time", "likert", ESPERA, None),
    (30, "ux_duration", "likert", CONCORDANCIA_CURTA, None),
    (31, "ux_clarity_info", "likert", QUALIDADE, None),
    (32, "ux_knowledge", "likert", QUALIDADE, None),
    (33, "ux_routing", "likert", QUALIDADE, None),
    (34, "ux_fairness", "likert", CONCORDANCIA_CURTA, None),
    (35, "ux_courtesy", "likert", SATISFACAO, None),
    (36, "ux_csat", "scale10", None, None),
    (37, "ux_expectations", "likert", EXPECTATIVAS, None),
    (38, "ux_nps", "nps", None, None),
]
SERVICE_COL_A = 6
CHECKLIST_COL_A = 9
CHECKLIST_COL_B = 8
CHECKLIST_COL_C = 9

COLMAP_B = [  # Questionário UX Online Certidão — 39 colunas, sem coluna de serviço
    (9, "ux_resolved", "sim_nao", None, None),
    (11, "ux_speed", "likert", RAPIDEZ, None),
    (12, "ux_channel_ease", "likert", FACILIDADE, "Presencial"),
    (13, "ux_channel_ease", "likert", FACILIDADE, "Online - Lisboa.pt"),
    (14, "ux_channel_ease", "likert", FACILIDADE, "Online - chat"),
    (15, "ux_channel_ease", "likert", FACILIDADE, "Telefone"),
    (16, "ux_findability", "likert", FACILIDADE, None),
    (17, "ux_language_clear", "likert", CONCORDANCIA_CURTA, None),
    (18, "ux_language_simple", "likert", CONCORDANCIA_CURTA, None),
    (19, "ux_response_clarity", "sim_nao", None, None),
    (23, "ux_ease", "likert", FACILIDADE, None),
    (24, "ux_scheduling", "agendamento", None, None),
    (26, "ux_scheduling_ease", "likert", FACILIDADE, None),
    (27, "ux_scheduling_date", "likert", FACILIDADE, None),
    (29, "ux_wait_time", "likert", ESPERA, None),
    (30, "ux_duration", "likert", CONCORDANCIA_CURTA, None),
    (31, "ux_clarity_info", "likert", QUALIDADE, None),
    (32, "ux_knowledge", "likert", QUALIDADE, None),
    (33, "ux_routing", "likert", QUALIDADE, None),
    (34, "ux_fairness", "likert", CONCORDANCIA_CURTA, None),
    (35, "ux_courtesy", "likert", SATISFACAO, None),
    (36, "ux_csat", "scale10", None, None),
    (37, "ux_expectations", "likert", EXPECTATIVAS, None),
    (38, "ux_nps", "nps", None, None),
]

COLMAP_C = [  # Questionário Recolha de Monstros — 40 colunas, cabeçalhos parciais corrompidos
    (10, "ux_resolved", "sim_nao", None, None),
    (12, "ux_speed", "likert", RAPIDEZ, None),
    (13, "ux_channel_ease", "likert", FACILIDADE, "Presencial"),
    (14, "ux_channel_ease", "likert", FACILIDADE, "Online - Na Minha Rua Lx"),
    (15, "ux_channel_ease", "likert", FACILIDADE, "Online - Lisboa.pt"),
    (16, "ux_channel_ease", "likert", FACILIDADE, "Online - chat"),
    (17, "ux_channel_ease", "likert", FACILIDADE, "Telefone"),
    (18, "ux_findability", "likert", FACILIDADE, None),
    (19, "ux_language_clear", "likert", CONCORDANCIA_CURTA, None),
    (20, "ux_language_simple", "likert", CONCORDANCIA_CURTA, None),
    (21, "ux_response_clarity", "sim_nao", None, None),
    (22, "ux_ease", "likert", FACILIDADE, None),
    (23, "ux_scheduling", "agendamento", None, None),
    (25, "ux_scheduling_ease", "likert", FACILIDADE, None),
    (26, "ux_scheduling_date", "likert", FACILIDADE, None),
    (27, "ux_wait_time", "likert", ESPERA, None),
    (28, "ux_duration", "likert", CONCORDANCIA_CURTA, None),
    (29, "ux_clarity_info", "likert", QUALIDADE, None),
    (30, "ux_knowledge", "likert", QUALIDADE, None),
    (31, "ux_routing", "likert", QUALIDADE, None),
    (32, "ux_fairness", "likert", CONCORDANCIA_CURTA, None),
    (33, "ux_courtesy", "likert", SATISFACAO, None),
    (37, "ux_csat", "scale10", None, None),
    (38, "ux_expectations", "likert", EXPECTATIVAS, None),
    (39, "ux_nps", "nps", None, None),
]

# Colunas intencionalmente não ingeridas em todos os ficheiros (sem indicador dedicado
# no catálogo, ou texto aberto/demografia) — mesmo critério de ingest_survey.py/ingest_iss_ux.py:
#   canais utilizados (resumo), Porquê/sugestões (texto aberto), cidadão vs empresa,
#   primeira vez, compreender decisão/estado/próximos passos, problemas no agendamento,
#   "realizou atendimento" (ficheiro B), assistente digital (ficheiro C).


def load_file(path, colmap, service_col=None, default_service=None, checklist_col=None):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    agg = defaultdict(lambda: {"codes": [], "cats": Counter()})  # key=(service_key, chan, etl, kind)
    unmapped = defaultdict(Counter)
    unmapped_service = Counter()
    n_by_service = Counter()
    unmapped_channel = Counter()
    channel_link_stats = Counter()  # ligado | vazio | multicanal | rotulo_desconhecido

    for row in rows:
        if all(v is None for v in row):
            continue
        if service_col:
            raw_serv = row[service_col - 1]
            sn = norm(raw_serv)
            if sn not in SERVICE_LABEL_MAP:
                unmapped_service[str(raw_serv).strip()] += 1
                continue
            service_key = SERVICE_LABEL_MAP[sn]
        else:
            service_key = default_service
        n_by_service[service_key] += 1

        resp_channel = None
        if checklist_col:
            raw_checklist = row[checklist_col - 1]
            before = sum(unmapped_channel.values())
            resp_channel = respondent_channel(raw_checklist, CHECKLIST_CHANNEL_MAP, unmapped_channel)
            if resp_channel is not None:
                channel_link_stats["ligado"] += 1
            elif raw_checklist in (None, ""):
                channel_link_stats["vazio"] += 1
            elif sum(unmapped_channel.values()) > before:
                channel_link_stats["rotulo_desconhecido"] += 1
            else:
                channel_link_stats["multicanal"] += 1

        for col, etl, kind, scale, chan in colmap:
            raw = row[col - 1]
            if raw in (None, ""):
                continue
            key = (service_key, chan, etl, kind)
            if kind == "sim_nao":
                nv = norm(raw)
                cat = SIM if nv == "sim" else NAO if nv in ("nao", "nao recebi resposta") else NA if nv == "nao aplicavel" else None
                if cat is None:
                    unmapped[etl][str(raw).strip()] += 1
                else:
                    agg[key]["cats"][cat] += 1
                    if chan is None and resp_channel is not None:
                        agg[(service_key, resp_channel, etl, kind)]["cats"][cat] += 1
            elif kind == "agendamento":
                nv = norm(raw)
                cat = SIM if nv == "sim" else NAO if nv == "nao" else NMT if nv in ("nao mas tentei",) else None
                if cat is None:
                    unmapped[etl][str(raw).strip()] += 1
                else:
                    agg[key]["cats"][cat] += 1
                    if chan is None and resp_channel is not None:
                        agg[(service_key, resp_channel, etl, kind)]["cats"][cat] += 1
            elif kind == "likert":
                nv = norm(raw)
                if nv in EXCLUDE:
                    continue
                code = scale.get(nv)
                if code is None:
                    unmapped[etl][str(raw).strip()] += 1
                else:
                    agg[key]["codes"].append(code)
                    if chan is not None:
                        # A UI escolhe o valor "atual" do card pela linha channel IS NULL —
                        # sem isto, ux_channel_ease só teria linhas por canal e nunca apareceria.
                        agg[(service_key, None, etl, kind)]["codes"].append(code)
                    elif resp_channel is not None:
                        # Ligação canal↔indicador: o inquirido só usou 1 canal, por isso esta
                        # resposta (hoje só agregada) também conta para esse canal específico.
                        agg[(service_key, resp_channel, etl, kind)]["codes"].append(code)
            elif kind in ("scale10", "nps"):
                try:
                    fv = float(raw)
                except Exception:
                    unmapped[etl][str(raw).strip()] += 1
                else:
                    agg[key]["codes"].append(fv)
                    if chan is None and resp_channel is not None:
                        agg[(service_key, resp_channel, etl, kind)]["codes"].append(fv)

    return dict(agg), unmapped, unmapped_service, n_by_service, unmapped_channel, channel_link_stats


def compute(kind, d):
    codes, cats = d["codes"], d["cats"]
    if kind in ("likert", "scale10"):
        n = len(codes)
        return (round(sum(codes) / n, 2) if n else None), None, n
    if kind == "nps":
        n = len(codes)
        if not n:
            return None, None, 0
        promo = sum(1 for v in codes if v >= 9)
        detr = sum(1 for v in codes if v <= 6)
        pas = n - promo - detr
        nps = round(100 * (promo - detr) / n, 1)
        return nps, {"Promotores": promo, "Passivos": pas, "Detratores": detr}, n
    if kind == "sim_nao":
        s, nn, na = cats[SIM], cats[NAO], cats[NA]
        den = s + nn
        val = round(100 * s / den, 1) if den else None
        cc = {SIM: s, NAO: nn}
        if na:
            cc[NA] = na
        return val, cc, s + nn + na
    if kind == "agendamento":
        s, nn, nmt = cats[SIM], cats[NAO], cats[NMT]
        return None, {SIM: s, NAO: nn, NMT: nmt}, s + nn + nmt
    return None, None, 0


def load_all():
    a = load_file(f"{BASE}/Questionário Presencial - Avaliação da experiência de utilização dos serviços.xlsx", COLMAP_A, service_col=SERVICE_COL_A, checklist_col=CHECKLIST_COL_A)
    b = load_file(f"{BASE}/Questionário UX Online -Certidão de Licença de Utilização - Avaliação da experiência de utilização do serviço.xlsx", COLMAP_B, default_service="certidao", checklist_col=CHECKLIST_COL_B)
    c = load_file(f"{BASE}/Questuinário UX Recolha de Monstros - Experiência de utilização do Serviço.xlsx", COLMAP_C, default_service="__monstros__", checklist_col=CHECKLIST_COL_C)
    return a, b, c


def q(s):
    return "'" + str(s).replace("'", "''") + "'"


def report():
    print("=" * 78)
    print("DRY-RUN — Ingestão dos 3 questionários UX novos da CML")
    print("=" * 78)
    for label, (agg, unmapped, unmapped_service, n_by_service, unmapped_channel, channel_link_stats) in zip(
        ["A: Presencial", "B: Online Certidão", "C: Recolha de Monstros"], load_all()
    ):
        print(f"\n── {label} ──")
        print(f"  respostas por serviço: {dict(n_by_service)}")
        print(f"  serviços não reconhecidos: {dict(unmapped_service) or '(nenhum)'}")
        print(f"  rótulos não mapeados: {({k: dict(v) for k, v in unmapped.items()}) or '(nenhum)'}")
        print(f"  ligação canal↔indicadores: {dict(channel_link_stats) or '(nenhum)'}")
        print(f"  canais não reconhecidos na checklist: {dict(unmapped_channel) or '(nenhum)'}")
        print(f"  grupos indicador×canal×serviço com dados: {len(agg)}")
        for (service_key, chan, etl, kind), d in sorted(agg.items(), key=lambda kv: str(kv[0])):
            value, cc, n = compute(kind, d)
            print(f"    {service_key:<18} {str(chan):<26} {etl:<20} n={n:<3} valor={value} {cc or ''}")


def emit_sql(out):
    results = load_all()
    for agg, unmapped, unmapped_service, _, _, _ in results:
        assert not unmapped, f"Há rótulos não mapeados, abortar: {dict(unmapped)}"
        assert not unmapped_service, f"Há serviços não reconhecidos, abortar: {dict(unmapped_service)}"

    L = ["-- Gerado por scripts/ingest_cml_ux.py — NÃO editar à mão.",
         "-- Fontes: docs/new/Questionários UX CML/ (3 ficheiros, 15 respostas no total).\n"]

    # Fundir os 3 ficheiros por chave ANTES de calcular — certidao aparece nos ficheiros
    # A e B; sem esta fusão, cada ficheiro gerava a sua própria linha para a MESMA chave
    # (service_id, indicator_id, year=2026, month=NULL, channel), duplicando dados na BD.
    merged = defaultdict(lambda: {"codes": [], "cats": Counter()})
    for agg, _, _, _, _, _ in results:
        for key, d in agg.items():
            merged[key]["codes"].extend(d["codes"])
            merged[key]["cats"].update(d["cats"])

    rows = []
    del_keys = []  # (service_id, etl, chstr) únicos, para o DELETE de pré-limpeza
    seen_del = set()
    for (service_key, chan, etl, kind), d in merged.items():
        value, cc, n = compute(kind, d)
        if service_key == "__monstros__":
            targets = [("monstros_pedido", True), ("monstros_servico", True)]
        else:
            targets = [(service_key, False)]
        # Linhas geradas pela ligação canal↔indicadores (chan definido mas o indicador não
        # é ux_channel_ease, i.e. o canal não vem de uma pergunta nativa por canal, é
        # atribuído a partir de "que canais utilizou") ficam marcadas is_provisional=TRUE.
        is_linked_channel = chan is not None and etl != "ux_channel_ease"
        for sk, provisional_target in targets:
            provisional = provisional_target or is_linked_channel
            service_id, _ = SERVICE_ID[sk]
            vstr = "NULL" if value is None else repr(float(value))
            cstr = "NULL" if cc is None else q(json.dumps(cc, ensure_ascii=False))
            chstr = "NULL" if chan is None else q(chan)
            rows.append(
                f"({q(service_id)}::uuid,{q(etl)},{chstr},{vstr},{cstr},{n},"
                f"{str(provisional).upper()},{q('cml_ux_questionarios_2026')})"
                )
            dk = (service_id, etl, chstr)
            if dk not in seen_del:
                seen_del.add(dk)
                del_keys.append(dk)

    if rows:
        # ON CONFLICT não é fiável aqui: month é sempre NULL nestas linhas, e o Postgres
        # trata NULL como distinto de NULL num índice único — duas linhas com month=NULL
        # NUNCA colidem entre si, mesmo com o resto da chave igual (id. para channel NULL).
        # Por isso apagamos primeiro qualquer linha equivalente (incl. o mock antigo da
        # migration 021) antes de inserir, usando IS NOT DISTINCT FROM para o canal.
        L.append(
            "DELETE FROM org_cml.measurements m\n"
            "USING (VALUES\n  " + ",\n  ".join(
                f"({q(service_id)}::uuid,{q(etl)},{chstr})" for service_id, etl, chstr in del_keys
            ) + "\n) AS v(service_id, etl, channel)\n"
            "JOIN public.indicators i ON i.etl_column_key = v.etl\n"
            "WHERE m.service_id = v.service_id AND m.indicator_id = i.id AND m.year = 2026\n"
            "  AND m.month IS NULL AND m.channel IS NOT DISTINCT FROM v.channel\n"
            "  AND m.geo_level IS NULL AND m.geo_name IS NULL;\n"
        )
        L.append(
            "INSERT INTO org_cml.measurements "
            "(service_id, indicator_id, year, month, channel, value, category_counts, "
            "total_respondentes, is_provisional, source_file)\n"
            "SELECT v.service_id, i.id, 2026, NULL, v.channel, v.value::numeric, v.cc::jsonb, v.resp, "
            "v.provisional, v.source\n"
            "FROM (VALUES\n  " + ",\n  ".join(rows) +
            "\n) AS v(service_id, etl, channel, value, cc, resp, provisional, source)\n"
            "JOIN public.indicators i ON i.etl_column_key = v.etl;"
        )

    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(L) + "\n")
    print(f"SQL escrito em {out}: {len(rows)} medições.")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--report", action="store_true")
    ap.add_argument("--sql")
    args = ap.parse_args()
    if args.sql:
        emit_sql(args.sql)
    else:
        report()


if __name__ == "__main__":
    main()
