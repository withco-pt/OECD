#!/usr/bin/env python3
"""
Ingestão de dados reais da ADC (Agência para o Desenvolvimento e Coesão) → org_adc.measurements.

Fontes (docs/ADC/):

  1. "Matriz_ Questionário de avaliação da experiência de utilização do Balcão dos
     Fundos (1-85).xlsx" — 85 linhas, 81 respostas utilizáveis (4 em branco/teste),
     repartidas pelos 2 serviços reais da ADC ("Apoio ao Balcão dos Fundos" e "Registo
     no Balcão dos Fundos" — nomes no ficheiro batem 1:1 com os nomes na BD).
     Layout de colunas idêntico ao "Matriz+LC" original (scripts/ingest_survey.py) —
     confirmado por validação exaustiva das 81 respostas: ZERO rótulos de escala fora
     dos dicionários já codificados em ingest_survey.py. Por isso reutilizam-se aqui os
     mesmos dicionários, apenas com um COLMAP novo (posições de coluna diferentes) e um
     conjunto de canais próprio (Telefone, E-mail, Suporte Escrito, Videochamada,
     ChatLive, IVO).

  2. "Recolha_Dados_Compliance_Matriz_ADC.xlsx" — autoavaliação de compliance
     respondida pela própria ADC (43 perguntas, sheet "2. Indicadores Compliance").
     Só as perguntas com "Tipo de indicador" = "Compliance" E que correspondem
     (por texto) a um dos 8 indicadores de compliance já no catálogo são ingeridas —
     as restantes ("Não-compliance", "Experiência de utilizador", ou sem indicador
     correspondente) ficam de fora, documentadas abaixo. Mapeamento manual, 1 a 1,
     revisto por descrição (ver COMPLIANCE_MAP).

     O ficheiro não distingue a qual dos 2 serviços cada resposta se aplica — é uma
     autoavaliação ao nível do "Balcão dos Fundos" como um todo. Por decisão (mesmo
     critério já usado para Monstros/CML): aplicada a AMBOS os serviços, marcada
     is_provisional=TRUE, com nota explícita da limitação.

Perguntas de compliance do ficheiro SEM correspondência no catálogo (não ingeridas):
  Interoperabilidade: autenticação federada (SSO)/Gov.pt redirect; Segurança de
  informação e dados abertos (5 perguntas RGPD/NIS2/dados.gov.pt); Envolvimento do
  utilizador (Livro Amarelo, desempenho do serviço, processo de conceção, barreiras
  linguísticas); "Se sim, tem periodicidade definida?..." (pergunta composta, sem
  equivalente direto). Ficam apenas documentadas nesta auditoria, não na BD.

Uso:
    python3 scripts/ingest_adc.py --report        # dry-run: valida e imprime
    python3 scripts/ingest_adc.py --sql out.sql   # gera SQL de ingestão

NADA é escrito na base de dados por este script.
"""
import argparse, json, unicodedata
from collections import defaultdict, Counter
from datetime import datetime

import openpyxl

BASE = "docs/ADC"

SERVICE_ID = {
    "apoio": "8f22b297-460f-4dfa-ad74-45aaed3a706d",       # Apoio ao Balcão dos Fundos
    "registo": "f20f03ae-8ad5-4f82-976d-12caa562f538",     # Registo no Balcão dos Fundos
}
SERVICE_LABEL_MAP = {
    "apoio ao balcao dos fundos": "apoio",
    "registo no balcao dos fundos": "registo",
}


def norm(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return " ".join(s.split())


# ── Dicionários — idênticos aos de scripts/ingest_survey.py (rótulos confirmados) ──
FACILIDADE = {"muito dificil": 1, "dificil": 2, "nem facil nem dificil": 3, "neutro": 3, "facil": 4, "muito facil": 5}
RAPIDEZ = {"muito lento": 1, "lento": 2, "nem lento nem rapido": 3, "rapido": 4, "muito rapido": 5}
CONCORDANCIA = {"discordo totalmente": 1, "discordo parcialmente": 2, "nao concordo nem discordo": 3, "concordo parcialmente": 4, "concordo totalmente": 5}
CONCORDANCIA_DUR = {"discordo totalmente": 1, "discordo": 2, "nao concordo nem discordo": 3, "concordo": 4, "concordo totalmente": 5}
QUALIDADE = {"muito mau": 1, "mau": 2, "nem mau nem bom": 3, "bom": 4, "muito bom": 5}
SATISFACAO = {"muito insatisfatorio": 1, "insatisfatorio": 2, "nem insatisfatorio nem satisfatorio": 3, "satisfatorio": 4, "muito satisfatorio": 5}
ESPERA = {"muito demorado/inaceitavel": 1, "demorado": 2, "tempo aceitavel": 3, "rapido": 4, "muito rapido": 5}
EXPECTATIVAS = {
    "nao correspondeu nada as expectativas": 1,
    "nao correspondeu em parte nem totalmente as expectativas": 2,
    "correspondeu em parte as expectativas": 3,
    "correspondeu totalmente as expectativas": 4,
    "superou as expectativas": 5,
}
EXCLUDE = {"nao utilizei este canal", "nao se aplica", "nao se aplica.", "prefiro nao responder", "ainda aguardo conclusao/resposta", "neutro"}

SIM = "Sim"; NAO = "Não"; NA = "Não aplicável"; NMT = "Não, mas tentei"

CHANNEL_LABEL = {13: "Telefone", 14: "E-mail", 15: "Suporte Escrito", 16: "Videochamada", 17: "ChatLive", 18: "IVO"}

SERVICE_COL = 8
DISTRITO_COL = 47

# kind: likert | scale10 | nps | sim_nao | agendamento
COLMAP = [
    (10, "ux_resolved", "sim_nao", None, None),
    (12, "ux_speed", "likert", RAPIDEZ, None),
    (13, "ux_channel_ease", "likert", FACILIDADE, "Telefone"),
    (14, "ux_channel_ease", "likert", FACILIDADE, "E-mail"),
    (15, "ux_channel_ease", "likert", FACILIDADE, "Suporte Escrito"),
    (16, "ux_channel_ease", "likert", FACILIDADE, "Videochamada"),
    (17, "ux_channel_ease", "likert", FACILIDADE, "ChatLive"),
    (18, "ux_channel_ease", "likert", FACILIDADE, "IVO"),
    (19, "ux_findability", "likert", FACILIDADE, None),
    (20, "ux_language_simple", "likert", CONCORDANCIA, None),
    (21, "ux_language_clear", "likert", CONCORDANCIA, None),
    (22, "ux_response_clarity", "sim_nao", None, None),
    (26, "ux_ease", "likert", FACILIDADE, None),
    (27, "ux_scheduling", "agendamento", None, None),
    (28, "ux_scheduling_ease", "likert", FACILIDADE, None),
    (29, "ux_scheduling_date", "likert", FACILIDADE, None),
    (31, "ux_wait_time", "likert", ESPERA, None),
    (32, "ux_duration", "likert", CONCORDANCIA_DUR, None),
    (33, "ux_clarity_info", "likert", QUALIDADE, None),
    (34, "ux_knowledge", "likert", QUALIDADE, None),
    (35, "ux_routing", "likert", QUALIDADE, None),
    (36, "ux_fairness", "likert", CONCORDANCIA, None),
    (37, "ux_courtesy", "likert", SATISFACAO, None),
    (38, "ux_csat", "scale10", None, None),
    (39, "ux_expectations", "likert", EXPECTATIVAS, None),
    (40, "ux_nps", "nps", None, None),
]
# Colunas intencionalmente não ingeridas: 9 (canais, resumo), 11/30/41 (texto aberto),
# 23-25 (sub-dimensões sem indicador dedicado), 42-46 (demografia).


def load_survey():
    path = f"{BASE}/Matriz_ Questionário de avaliação da experiência de utilização do Balcão dos Fundos\xa0(1-111).xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]

    agg = defaultdict(lambda: {"codes": [], "cats": Counter()})       # key=(service_key, chan, etl, kind)
    agg_geo = defaultdict(lambda: {"codes": [], "cats": Counter()})   # key=(service_key, distrito, etl, kind)
    pop = Counter()
    unmapped = defaultdict(Counter)
    unmapped_service = Counter()
    n_by_service = Counter()

    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        raw_serv = row[SERVICE_COL - 1]
        if raw_serv is None:
            continue
        sn = norm(raw_serv)
        if sn not in SERVICE_LABEL_MAP:
            unmapped_service[str(raw_serv).strip()] += 1
            continue
        service_key = SERVICE_LABEL_MAP[sn]
        n_by_service[service_key] += 1
        pop[service_key] += 1
        distrito = row[DISTRITO_COL - 1]
        distrito = str(distrito).strip() if distrito not in (None, "") else None
        if distrito and norm(distrito) == "prefiro nao responder":
            distrito = None

        for col, etl, kind, scale, chan in COLMAP:
            raw = row[col - 1]
            if raw in (None, ""):
                continue
            key = (service_key, chan, etl, kind)
            geo_key = (service_key, distrito, etl, kind) if distrito else None

            if kind == "sim_nao":
                nv = norm(raw)
                cat = SIM if nv == "sim" else NAO if nv in ("nao", "nao recebi resposta") else NA if nv == "nao aplicavel" else None
                if cat is None:
                    unmapped[etl][str(raw).strip()] += 1
                else:
                    agg[key]["cats"][cat] += 1
                    if geo_key:
                        agg_geo[geo_key]["cats"][cat] += 1
            elif kind == "agendamento":
                nv = norm(raw)
                cat = SIM if nv == "sim" else NAO if nv == "nao" else NMT if nv in ("nao mas tentei", "nao, mas tentei") else None
                if cat is None:
                    unmapped[etl][str(raw).strip()] += 1
                else:
                    agg[key]["cats"][cat] += 1
                    if geo_key:
                        agg_geo[geo_key]["cats"][cat] += 1
            elif kind == "likert":
                nv = norm(raw)
                if nv in EXCLUDE:
                    continue
                code = scale.get(nv)
                if code is None:
                    unmapped[etl][str(raw).strip()] += 1
                else:
                    agg[key]["codes"].append(code)
                    if geo_key:
                        agg_geo[geo_key]["codes"].append(code)
                    if chan is not None:
                        agg[(service_key, None, etl, kind)]["codes"].append(code)
            elif kind in ("scale10", "nps"):
                try:
                    fv = float(raw)
                except Exception:
                    unmapped[etl][str(raw).strip()] += 1
                else:
                    agg[key]["codes"].append(fv)
                    if geo_key:
                        agg_geo[geo_key]["codes"].append(fv)

    return dict(agg), dict(agg_geo), pop, unmapped, unmapped_service, n_by_service


def compute(kind, d):
    codes, cats = d["codes"], d["cats"]
    if kind in ("likert", "scale10"):
        n = len(codes)
        return (round(sum(codes) / n, 2) if n else None), None, n
    if kind == "nps":
        n = len(codes)
        if not n:
            return None, None, 0
        promo = sum(1 for v in codes if v >= 9)
        detr = sum(1 for v in codes if v <= 6)
        pas = n - promo - detr
        nps = round(100 * (promo - detr) / n, 1)
        return nps, {"Promotores": promo, "Passivos": pas, "Detratores": detr}, n
    if kind == "sim_nao":
        s, nn, na = cats[SIM], cats[NAO], cats[NA]
        den = s + nn
        val = round(100 * s / den, 1) if den else None
        cc = {SIM: s, NAO: nn}
        if na:
            cc[NA] = na
        return val, cc, s + nn + na
    if kind == "agendamento":
        s, nn, nmt = cats[SIM], cats[NAO], cats[NMT]
        return None, {SIM: s, NAO: nn, NMT: nmt}, s + nn + nmt
    return None, None, 0


# ── Compliance — mapeamento manual (ver docstring) ──────────────────────────
# (indicator_id, texto da pergunta no ficheiro ADC a localizar por match exato normalizado)
COMPLIANCE_MAP = [
    ("37732d56-b185-47ca-a827-71a327d524cd", "Utiliza o Ágora Design System?"),
    ("79d2668b-dde1-496b-a6c7-b9db32ce0f0c", "O formulário para solicitação do serviço apresenta campos pré-preenchidos com dados disponíveis na entidade?"),
    ("88467533-a43c-4e47-8293-62320cf47b2d", "O serviço está disponivel na LC/Espaço Cidadão?"),
    ("4c03f8e5-bbf4-4b3e-969d-155f3905ac0d", "O serviço está disponível em inglês?"),
    ("f6662b5a-3efe-4027-9d0b-d9516bddbf1d", "O serviço público está disponível num regime de atendimento omnicanal (ex. canais presencial, web/app e telefónico )?"),
    ("4850d9fb-1934-4759-b084-27523d453856", "São solicitados documentos ao cidadão que estão disponíveis noutras entidades da Administração Pública?"),
    ("0951ab7e-d684-4cf6-be6a-48599556fa79", "São recolhidos dados sobre a satisfação do utilizador sobre a qualidade do serviço?"),
]


def load_compliance():
    path = f"{BASE}/Recolha_Dados_Compliance_Matriz_ADC.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["2. Indicadores Compliance"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    by_question = {}
    for row in rows:
        indicador, questao, tipo, resposta = row[1], row[2], row[3], row[7]
        by_question[norm(indicador)] = (tipo, resposta, row[8])

    results = []
    unmatched = []
    for indicator_id, question_text in COMPLIANCE_MAP:
        key = norm(question_text)
        if key not in by_question:
            unmatched.append(question_text)
            continue
        tipo, resposta, comentario = by_question[key]
        nv = norm(resposta)
        cat = SIM if nv == "sim" else NAO if nv == "nao" else NA if nv in ("na", "nao aplicavel") else None
        if cat is None:
            unmatched.append(f"{question_text} (resposta não reconhecida: {resposta!r})")
            continue
        results.append({
            "indicator_id": indicator_id, "question": question_text,
            "resposta": resposta, "cat": cat, "comentario": comentario,
        })
    return results, unmatched


def q(s):
    return "'" + str(s).replace("'", "''") + "'"


def report():
    print("=" * 78)
    print("DRY-RUN — Ingestão de dados reais da ADC")
    print("=" * 78)

    agg, agg_geo, pop, unmapped, unmapped_service, n_by_service = load_survey()
    print(f"\n[1] Questionário Balcão dos Fundos — respostas por serviço: {dict(n_by_service)}")
    print(f"  serviços não reconhecidos: {dict(unmapped_service) or '(nenhum)'}")
    print(f"  rótulos não mapeados: {({k: dict(v) for k, v in unmapped.items()}) or '(nenhum)'}")
    print(f"  grupos indicador×canal×serviço com dados: {len(agg)}")
    print(f"  grupos por distrito: {len(agg_geo)}")

    results, unmatched = load_compliance()
    print(f"\n[2] Compliance ADC — {len(results)}/{len(COMPLIANCE_MAP)} perguntas mapeadas ao catálogo")
    for r in results:
        print(f"  {r['question'][:70]:<70} -> {r['resposta']!r} (cat={r['cat']})" + (f"  [nota: {r['comentario']}]" if r['comentario'] else ""))
    if unmatched:
        print(f"  NÃO encontradas/reconhecidas: {unmatched}")


def emit_sql(out):
    agg, agg_geo, pop, unmapped, unmapped_service, _ = load_survey()
    assert not unmapped, f"Há rótulos não mapeados, abortar: {dict(unmapped)}"
    assert not unmapped_service, f"Há serviços não reconhecidos, abortar: {dict(unmapped_service)}"
    results, unmatched = load_compliance()

    L = ["-- Gerado por scripts/ingest_adc.py — NÃO editar à mão.",
         "-- Fontes: docs/ADC/ (questionário UX + autoavaliação de compliance da ADC).\n"]

    # ── UX: linhas por canal/distrito ──
    # DELETE de pré-limpeza só é necessário para as linhas agregadas (channel/geo NULL),
    # que são as únicas que podem colidir com o mock antigo da migration 021 (que nunca
    # teve granularidade por canal ou distrito). As linhas por canal/distrito são todas
    # novas — nenhum risco de duplicação, não precisam de chave de limpeza.
    ux_rows = []
    del_keys = []
    seen_del = set()
    for (service_key, chan, etl, kind), d in agg.items():
        value, cc, n = compute(kind, d)
        service_id = SERVICE_ID[service_key]
        vstr = "NULL" if value is None else repr(float(value))
        cstr = "NULL" if cc is None else q(json.dumps(cc, ensure_ascii=False))
        chstr = "NULL" if chan is None else q(chan)
        ux_rows.append(f"({q(service_id)}::uuid,{q(etl)},{chstr},NULL,NULL,{vstr},{cstr},{n},{q('adc_ux_questionario_2026')})")
        if chan is None:
            dk = (service_id, etl)
            if dk not in seen_del:
                seen_del.add(dk); del_keys.append(dk)
    for (service_key, distrito, etl, kind), d in agg_geo.items():
        value, cc, n = compute(kind, d)
        service_id = SERVICE_ID[service_key]
        vstr = "NULL" if value is None else repr(float(value))
        cstr = "NULL" if cc is None else q(json.dumps(cc, ensure_ascii=False))
        ux_rows.append(f"({q(service_id)}::uuid,{q(etl)},NULL,{q('distrito')},{q(distrito)},{vstr},{cstr},{n},{q('adc_ux_questionario_2026')})")

    if del_keys:
        L.append(
            "DELETE FROM org_adc.measurements m\n"
            "USING (VALUES\n  " + ",\n  ".join(
                f"({q(sid)}::uuid,{q(etl)})" for sid, etl in del_keys
            ) + "\n) AS v(service_id, etl)\n"
            "JOIN public.indicators i ON i.etl_column_key = v.etl\n"
            "WHERE m.service_id = v.service_id AND m.indicator_id = i.id AND m.year = 2026\n"
            "  AND m.month IS NULL AND m.channel IS NULL AND m.geo_level IS NULL AND m.geo_name IS NULL;\n"
        )
    if ux_rows:
        L.append(
            "INSERT INTO org_adc.measurements "
            "(service_id, indicator_id, year, month, channel, geo_level, geo_name, value, category_counts, "
            "total_respondentes, source_file)\n"
            "SELECT v.service_id, i.id, 2026, NULL, v.channel, v.geo_level, v.geo_name, v.value::numeric, "
            "v.cc::jsonb, v.resp, v.source\n"
            "FROM (VALUES\n  " + ",\n  ".join(ux_rows) +
            "\n) AS v(service_id, etl, channel, geo_level, geo_name, value, cc, resp, source)\n"
            "JOIN public.indicators i ON i.etl_column_key = v.etl;\n"
        )

    # ── Compliance: aplicado a ambos os serviços (autoavaliação entity-wide) ──
    if results:
        comp_values = []
        comp_del_keys = []
        for r in results:
            for service_key in ("apoio", "registo"):
                service_id = SERVICE_ID[service_key]
                s_count = 1 if r["cat"] == SIM else 0
                n_count = 1 if r["cat"] == NAO else 0
                na_count = 1 if r["cat"] == NA else 0
                den = s_count + n_count
                value = "NULL" if den == 0 else repr(round(100 * s_count / den, 1))
                cc = {"Sim": s_count, "Não": n_count}
                if na_count:
                    cc["Não aplicável"] = na_count
                comp_values.append(
                    f"({q(service_id)}::uuid,{q(r['indicator_id'])}::uuid,2026,NULL,{value},"
                    f"{q(json.dumps(cc, ensure_ascii=False))},TRUE,{q('adc_compliance_autoavaliacao_2026')})"
                )
                comp_del_keys.append((service_id, r["indicator_id"]))

        L.append(
            "DELETE FROM org_adc.measurements m\n"
            "USING (VALUES\n  " + ",\n  ".join(
                f"({q(sid)}::uuid,{q(iid)}::uuid)" for sid, iid in comp_del_keys
            ) + "\n) AS v(service_id, indicator_id)\n"
            "WHERE m.service_id = v.service_id AND m.indicator_id = v.indicator_id AND m.year = 2026\n"
            "  AND m.month IS NULL AND m.channel IS NULL AND m.geo_level IS NULL AND m.geo_name IS NULL;\n"
        )
        L.append(
            "INSERT INTO org_adc.measurements "
            "(service_id, indicator_id, year, month, value, category_counts, is_provisional, source_file)\n"
            "VALUES\n  " + ",\n  ".join(comp_values) + ";"
        )

    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(L) + "\n")
    print(f"SQL escrito em {out}: {len(ux_rows)} linhas UX + {len(results)*2 if results else 0} linhas compliance.")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--report", action="store_true")
    ap.add_argument("--sql")
    args = ap.parse_args()
    if args.sql:
        emit_sql(args.sql)
    else:
        report()


if __name__ == "__main__":
    main()
